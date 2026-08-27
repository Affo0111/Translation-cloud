#!/usr/bin/env python3
"""
表格翻译系统 - Streamlit Web 界面
核心引擎：translator.py（load_template + apply_rules）
格式保留：openpyxl 直接操作原始文件副本

用法：
    streamlit run translator_streamlit.py
"""

import io
import os
import tempfile
from contextlib import contextmanager
from pathlib import Path
from typing import Optional

import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import openpyxl
import logging
import re
import shutil

# ── 抑制 translator.py 的 INFO 日志（避免 Streamlit 显示为报错）──
logging.getLogger("translator").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)

# ── 从 translator.py 导入核心函数 ────────────────────────────────
from translator import (
    _find_column,
    parse_rule,
    apply_rules,
    load_template,
    process_orders_preserve_format,
    _update_xlsx_cells_lightweight,
    _col_idx_to_letter,
)

# ════════════════════════════════════════════════════════════════════════════
# 规则说明弹窗（A / B 系统共用）
# ════════════════════════════════════════════════════════════════════════════

_A_SYSTEM_RULES_MD = r"""
### 作用
根据 SKU 规则，把订单中的「定制项」列（K 列）翻译成工厂能看懂的格式，并可按条件改 SKU、删行、增行。

### 操作符（按执行顺序）
| 符号 | 名称 | 写法 | 示例 |
|------|------|------|------|
| `^` | SKU 变化 | `^[条件] , 后缀` | `^[尺寸:L] , -1` → 满足条件时 SKU 末尾加 `-1` |
| `!` | 删除 | `![条件]` | `![无:XYZ]` → 删除「无:XYZ」行 |
| `=` | 翻译 | `=[源]=[目标]` | `=[Name]=[名字]` 改键名；`=[尺寸:L]=[尺寸:大号]` 改键值 |
| `++` | 指定位置插入 | `[位置] ++ [内容]` | `[前面有云:白云] ++ [背后有云:白云]` |
| `+` | 行末追加 | `+[内容]` | `+[注意:加急]` |

### 修饰符
| 符号 | 含义 | 示例 |
|------|------|------|
| `[]` | 范围/条件 | `[尺寸:L]` |
| `:` | 键值分隔 | `键:值` |
| `&` | 逻辑与 | `[尺寸:L]&[颜色:红]` |
| `\|` | 逻辑或 | `[尺寸:L\|尺寸:M]` |
| `;` | 子规则分隔 | 多条规则用 `;` 隔开 |

### 优先级
执行顺序：`^` → `!` → `=` → `++` → `+`；同一规则内先 `[]` 再 `&` 再 `\|`。
"""

_B_SYSTEM_RULES_MD = r"""
### 作用
根据订单「临时列(需删除)」（J 列）生成「callie定制项」（AZ 列），并按 SKU 从翻译模板「calie商品ID和商品版本」列自动填充：
- **AW 列「参考callie站点」** → 只有该 SKU 在「calie商品ID和商品版本」列配了 callie 商品信息时才填 `callie`
- **AX 列「calie商品ID」** → 来自「calie商品ID和商品版本」列
- **AY 列「callie商品版本」** → 来自「calie商品ID和商品版本」列

> 列位置以原始订单 `pet8.17.xls` 为准：AW=参考callie站点、AX=calie商品ID、AY=callie商品版本、AZ=callie定制项。

### 翻译模板列布局（方案 A）
| 列 | 内容 |
|------|------|
| sku 列 | 产品 SKU（Pet 在 B 列、Supply 在 C 列，按表头名自动识别） |
| **标准化规则**（原 D 列） | J 列标准化前置规则，详见下方「标准化规则（阶段 0）」 |
| **callie定制项翻译规则**（原 E 列） | 下方语法表，编译生成 AZ |
| 【临时列(需删除)】示例（原 F 列） | J 示例 |
| 【callie定制项】示例（原 G 列） | AZ 示例 |
| calie商品ID和商品版本（原 H 列） | 商品信息，填充 AW/AX/AY |

### 语法（callie定制项翻译规则 列）
| 符号 | 名称 | 写法 | 示例 |
|------|------|------|------|
| `@template` | 模板切换 | `@template: 字段名` 或 `@template: 字段名 = 缺省值` | `@template: Gender` → 按 Gender 值切换男女模板；`@template: Gender = Woman` → 订单 J 无 Gender 时默认用女性模板 |
| `+[字段:默认值]` | 固定字段 | `+[Facial Expression:Open Eyes]` | 自动查属性表补 ID 输出 |
| `+[ID\|字段名:\|值]` | 固定行 | `+[104\|Hijab Option:\|without Hijab]` | 直接输出固定 gid 行 |
| `+[ID\|字段:\|[JKey]]` | 带占位符固定行 | `+[170\|Color:\|[Color]]` | 用 J 列 Color 值替换占位符 |
| `[字段]` | 字段映射 | `[Hair Color]` | 把 J 列 Hair Color 值映射到属性表对应 ID |
| `? [条件], +[...]` | 条件固定行 | `? [Gender:Woman],+[104\|...]` | 满足条件才输出 |
| `? [A] & [B], ...` | 多条件 AND | `? [Gender:Woman]&[Age:Young],...` | 同时满足才执行 |
| `[字段1:值]=[ID]` | 条件路由 | `? [Gender:Woman],[Age:Young\|Senior]=[128\|126]` | 多值一一映射 |
| `!key` | 忽略字段 | `!Style` | 不输出该字段的隐式映射 |
| `#` | 注释 | `# 这是注释` | 不解析 |

### 隐式映射
只要 J 列字段在属性表中能找到，**即使 callie定制项翻译规则列没写 `[字段]`**，引擎也会自动补 ID 输出。只有专属字段或需要特殊覆盖时才用手写规则。

### 标准化规则（阶段 0，写入「标准化规则」列）
J 列（临时列）格式常不一致（组合字段、字段名多余、重复字段、空格/大小写杂质），会在「规则执行之前」先做标准化清洗，提升匹配准确率。

**语法**（单单元格内换行写多条，每条以 `rule|` 开头，行尾可带 `;`）：
```
rule|原始行|目标行1|目标行2|...;
```
- **原始行**：J 列中某行的整行精确内容（匹配前自动 `strip()`，所以首尾空格不影响）；
- **目标行 1..N**：标准化后展开成的一到多行；目标行留空（`rule|原始行|`）表示**删除该行**；
- 原始行/目标行**不得含 `|`**（用作分隔符）；
- 执行：**单轮、非链式、按书写顺序**；一条原始行命中第一条能匹配的规则后即替换，不再参与后续匹配；未匹配行原样保留。

**通配符**（原始行含 `*` 时启用）：
```
rule|Hairstyle:*-Hairstyle-*|Hair Color:$1|Hairstyle:Hairstyle-$2;
```
- `*` 匹配任意内容（贪婪），目标行可用 `$1`、`$2`… 引用对应的 `*` 捕获组；
- 上例把合并字段 `Hairstyle:Deepest Blue Black-Hairstyle-2` 拆成 `Hair Color:Deepest Blue Black` + `Hairstyle:Hairstyle-2`，**一条规则覆盖所有发色/发型**，无需逐个枚举；
- 精确规则（无 `*`）优先于通配规则匹配；原始行含 `*` 才会走通配。

**完整示例（PL251368）**
```
rule|Skin Tone:Young 1|Age:Young|Skin Tone:1;
rule|Beard Color - Style:Dark Brown-8|Beard Color:Dark Brown|Beard Style:8;
```
第一行把 `Skin Tone:Young 1` 拆成 `Age:Young` + `Skin Tone:1`；第二行把合并字段 `Beard Color - Style:Dark Brown-8` 拆成两个独立字段。标准化后的 J 再交给上方翻译规则处理。

### 单性别产品（订单 J 无 Gender 字段）
有些产品属性表里有男女两套模板，但**只做其中一个性别**（如只做女性），此时订单 J 列不会出现 `Gender` 字段。直接写 `@template: Gender` 会因读不到值而失效（跨模板字段输出字面 `template|`）。

**正确写法**：`@template` 加缺省值 + `+[Gender:...]` 固定字段两行配合：
```text
@template: Gender = Woman     # 等号后=缺省模板：J 无 Gender 时默认按女性解析 gid
+[Gender:Woman]               # 固定字段：①输出 100|Gender:|Woman 行 ②让下方 ? [Gender:Woman] 守卫读到值
? [Gender:Woman],[Hair Color:9色]=[108~116];   # 发色→发型路由（守卫靠上面固定字段注入的 Gender）
[Hair Color];
```
> 两行缺一不可：`@template = Woman` 只解决「字段 gid 解析」；`? [Gender:Woman]` 路由守卫仍需 `+[Gender:Woman]` 把值塞进条件才能触发。若将来订单 J 里真的出现 `Gender` 字段，会自动覆盖缺省值、走对应模板。

### 完整示例（PL251368，每行带 `#` 注释解释含义）
```text
@template: Gender                    # 按 Gender 切换男/女属性表模板，决定各字段用哪套 gid

# —— 样式与背景：条件固定行，命中即输出固定 gid 行（不是「字段=值」映射）——
? [Style:Vertical],+[1|:10001|];    # 客人选 Vertical  → 输出 1|…|10001
? [Style:Horizontal],+[1|:10002|];  # 客人选 Horizontal → 输出 1|…|10002
? [Background Style:Pencil],+[165|:1650001|];   # 选 Pencil 背景 → 165|…|1650001
? [Background Style:Crayon],+[165|:1650002|];   # 选 Crayon 背景 → 165|…|1650002
? [Background Style:Crayon],+[170|Color:|[Color]];  # Crayon 时还要带颜色：取 J 列 Color 值填入 170|Color:|<值>

# —— 年龄路由：同一字段在多 gid 间切换，属性表没有「值→gid」标记，必须手写 ——
? [Gender:Woman], [Age:Young|Age:Senior]=[128|126];  # 女+年轻→皮肤 128，女+年长→皮肤 126
? [Gender:Woman], [Age:Young|Age:Senior]=[125|124];  # 女+年轻→眼睛 125，女+年长→眼睛 124
? [Gender:Man],   [Age:Young|Age:Senior]=[163|164];  # 男+年轻→皮肤 163，男+年长→皮肤 164
? [Gender:Man],   [Age:Young|Age:Senior]=[159|160];  # 男+年轻→眼睛 159，男+年长→眼睛 160

[Hair Color];   # 裸字段映射：自动把 J 列 Hair Color 值映射到对应 gid（男/女套由 @template 决定）

# —— 发色路由：不同性别发色对应不同发型 gid（同属「模板内多 gid」，必须手写）——
? [Gender:Woman], [Hair Color:Black|Hair Color:Dark Brown|Hair Color:Light Blonde|Hair Color:Silver]=[107|108|110|115];  # 女：黑→107 深棕→108 浅金→110 银→115
? [Gender:Man],   [Hair Color:Black|Hair Color:Dark Brown|Hair Color:Light Blonde|Hair Color:Silver]=[133|134|136|140];  # 男：黑→133 深棕→134 浅金→136 银→140

+[Facial Expression:Open Eyes];   # 固定字段：所有人统一 Open Eyes（自动查属性表补 ID，不分男女）
? [Gender:Woman],+[104|Hijab Option:|without Hijab];   # 专属字段：仅女性输出 Hijab Option（男无此字段，故带 Gender 守卫）
```

### J 列字段拆分规则（以 PL251368 为例）
部分 SKU 在亚马逊 J 列会把「颜色 + 样式」合并成一个字段（如 `Beard Color - Style:Dark Brown-8`），但规则是按 **`Beard Color` 和 `Beard Style` 两个独立字段**设计的，引擎无法直接识别合并字段。

**手动拆分方式**（在 J 列把该合并字段拆成两行）：
| 原始 J 字段 | 拆成 |
|------|------|
| `Beard Color - Style:Dark Brown-8` | `Beard Color:Dark Brown` + `Beard Style:8` |
| `Beard Color - Style:Black-3` | `Beard Color:Black` + `Beard Style:3` |

即：**减号前 = 颜色 → `Beard Color:`；减号后 = 样式号 → `Beard Style:`**。

⚠️ 注意：
1. **两个都得拆**：只拆颜色不拆样式会丢样式号（只出 `142|Beard Color:|xxx` 而无样式 gid）。
2. **颜色必须是规则已覆盖的 4 个值**：`Black / Dark Brown / Light Blonde / Silver`。若出现规则没列的颜色（如 `Brown`、`Blonde`、`Ginger`），路由不触发，只会出颜色行而无样式行——需先在 callie定制项翻译规则列里补上对应颜色。
3. 漏拆的合并字段会在导出报告里报 `部分未匹配: Beard Color - Style`，拆干净后即消失。

> `#` 开头的行为注释，引擎解析时会自动忽略；上面示例里的注释只是给人看的说明，你抄规则时可以整行删掉。
"""


@st.dialog("📖 A 系统规则说明", width="large")
def _show_rules_a():
    st.markdown(_A_SYSTEM_RULES_MD)


@st.dialog("📖 B 系统规则说明", width="large")
def _show_rules_b():
    st.markdown(_B_SYSTEM_RULES_MD)


def _list_registered_callie_skus():
    """返回已入库的属性表 SKU 列表（排除演示配置 CZYDEMO）。"""
    if not os.path.isdir(CALLIE_SKU_DIR):
        return []
    return sorted(
        d for d in os.listdir(CALLIE_SKU_DIR)
        if d != "CZYDEMO"
        and os.path.isdir(os.path.join(CALLIE_SKU_DIR, d))
        and os.path.isfile(os.path.join(CALLIE_SKU_DIR, d, "attribute_config.json"))
    )


@st.dialog("📚 已入库属性表", width="small")
def _show_registered_skus():
    skus = _list_registered_callie_skus()
    if skus:
        st.markdown("以下 SKU 的属性表已入库，B 系统导出时会自动复用：")
        for sku in skus:
            st.markdown(f"- `{sku}`")
    else:
        st.info("暂无已入库属性表")


# ── 页面配置 ────────────────────────────────────────────────────
st.set_page_config(
    page_title="订单翻译系统",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #4f46e5 0%, #7c3aed 100%);
        padding: 1.2rem 2rem; border-radius: 12px; color: white; margin-bottom: 1rem;
        position: relative;
    }
    .main-header h1 { margin: 0; font-size: 1.4rem; }
    .main-header p { margin: 4px 0 0; opacity: 0.85; font-size: 0.85rem; }
    /* 步骤卡片：给 st.container(border=True) 的边框包装器加样式 */
    [data-testid="stVerticalBlockBorderWrapper"] {
        background: #ffffff; border: 1px solid #e2e8f0 !important;
        border-radius: 10px; box-shadow: 0 1px 3px rgba(0,0,0,0.06);
    }
    [data-testid="stVerticalBlockBorderWrapper"] > div { padding: 0.5rem 1rem 0.5rem 1rem !important; }
    [data-testid="stVerticalBlockBorderWrapper"] h4 {
        font-size: 0.95rem; margin: 0 0 0.5rem; color: #4f46e5;
    }
    section[data-testid="stFileUploader"] { margin-top: 0 !important; }
    .stat-box {
        background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px;
        padding: 0.5rem; text-align: center; margin-bottom: 0.4rem;
    }
    .stat-box .number { font-size: 1.3rem; font-weight: 700; color: #4f46e5; }
    .stat-box .label { font-size: 0.7rem; color: #64748b; }
    .saved-badge {
        display: inline-block; background: #dbeafe; color: #1e40af;
        padding: 3px 10px; border-radius: 20px; font-size: 0.78rem; font-weight: 600; margin: 4px 0;
    }
    .col-detect { font-size: 0.75rem; color: #059669; margin: 2px 0; }
    section[data-testid="stFileUploader"] { padding: 0 !important; }
    div[data-testid="stFileUploaderDropzone"] { padding: 0.5rem !important; font-size: 0.8rem !important; }
    div[data-testid="stFileUploaderDropzone"] small { font-size: 0.7rem !important; }
    .rules-btn {
        text-align: right;
    }
</style>
""", unsafe_allow_html=True)

# A 系统 header + 右上角规则说明按钮
_h_left, _h_right = st.columns([6, 1])
with _h_left:
    st.markdown("""
    <div class="main-header">
        <h1>🛠️ 定制项翻译</h1>
    </div>
    """, unsafe_allow_html=True)
with _h_right:
    st.markdown('<div style="height:18px"></div>', unsafe_allow_html=True)
    if st.button("📖 规则说明", key="btn_rules_a", use_container_width=True):
        _show_rules_a()

# ── 会话状态 ────────────────────────────────────────────────────
for key, default in [
    ("template_df", None), ("template_name", ""), ("template_saved", False),
    ("template_rules", None), ("parse_errors_list", []),
    ("orders_df", None), ("orders_name", ""), ("orders_saved", False),
    ("orders_wb", None), ("orders_path", None),
    ("result_path", None), ("stat_modified", 0), ("stat_skipped", 0), ("stat_ac_as", 0),
    ("stat_callie", 0), ("callie_cfgs", {}), ("template_path", None),
    ("callie_output", []), ("callie_unmatched", []), ("callie_section_count", 0),
    ("exec_errors_list", []), ("processing_done", False),
    ("csv_template_text", ""), ("csv_orders_text", ""),
    ("corrector_input", ""), ("corrector_output", ""),
    ("corrector_check_lines", []), ("corrector_has_run", False),
    ("generator_before", ""), ("generator_after", ""), ("generator_rules", ""),
    ("auto_export_msg", None), ("auto_orders_list", []),
    # B 系统状态
    ("callie_order_df", None), ("callie_order_path", None), ("callie_order_name", ""),
    ("callie_order_saved", False),
    ("callie_template_path", None), ("callie_template_name", ""),
    ("callie_template_saved", False),
    ("callie_result_path", None), ("callie_errors", []), ("callie_cfg_summary", {}),
    ("callie_reused_skus", []), ("callie_product_map_len", 0),
]:
    if key not in st.session_state:
        st.session_state[key] = default


def _cleanup(path: Optional[str]):
    if path:
        try: Path(path).unlink(missing_ok=True)
        except Exception: pass


@contextmanager
def _step_card(title: str):
    """步骤卡片容器：带边框圆角背景，避免 st.markdown('<div>') 产生空 <p> 占位。"""
    with st.container(border=True):
        st.markdown(f"#### {title}")
        yield


def _ensure_callie_loaded():
    """加载 callie 各 SKU 配置（存系统），存入会话状态。独立功能，失败不影响翻译。"""
    try:
        from callie_generator import load_sku_configs as _lsc
        st.session_state.callie_cfgs = _lsc(CALLIE_SKU_DIR) if os.path.isdir(CALLIE_SKU_DIR) else {}
        if st.session_state.callie_cfgs:
            logger.info(f"callie 已就绪: {len(st.session_state.callie_cfgs)} 份 SKU 配置")
    except Exception as e:
        logger.warning(f"callie 配置加载失败（不影响翻译）: {e}")
        st.session_state.callie_cfgs = {}


def _add_callie_alias(sku, field, raw, standard):
    """自学习：把新写法加入该 SKU 的 rule.json（仅 Style 字段支持别名归一，其余字段跳过）。"""
    if not standard:
        return
    if field != "Style":
        logger.warning(f"自学习暂仅支持 Style 字段，忽略 {sku}/{field}")
        return
    import json
    path = os.path.join(CALLIE_SKU_DIR, sku, "rule.json")
    if not os.path.isfile(path):
        return
    with open(path, encoding="utf-8") as f:
        rule = json.load(f)
    sf = rule.get("style_field") or {}
    aliases = sf.get("aliases") or {}
    aliases[raw] = standard
    sf["aliases"] = aliases
    rule["style_field"] = sf
    with open(path, "w", encoding="utf-8") as f:
        json.dump(rule, f, ensure_ascii=False, indent=2)
    _ensure_callie_loaded()


def _parse_callie_upload(uploaded):
    """把上传的 [SKU, 临时列] 表格规整为统一两列。"""
    if uploaded.name.lower().endswith(".csv"):
        df = pd.read_csv(uploaded, dtype=str)
    else:
        df = pd.read_excel(uploaded, dtype=str)
    cols = list(df.columns)
    if len(cols) < 2:
        raise ValueError("上传表格至少需要两列（SKU、临时列）")
    sku_col = _find_column(df, ["SKU", "sku", "商品编码"]) or cols[0]
    j_col = _find_column(df, ["临时列(需删除)", "临时列", "J列", "J", "定制项", "定制"]) or cols[1]
    return df[[sku_col, j_col]].rename(columns={sku_col: "SKU", j_col: "临时列(需删除)"})


def _convert_to_temp(order_path):
    """翻译单个订单文件，写入临时结果文件，返回 (out_path, modified, skipped, error_count, ac_as_count)。"""
    out_path = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
    modified, skipped, error_count, ac_as_count = process_orders_preserve_format(
        order_path, out_path, st.session_state.template_rules)
    return out_path, modified, skipped, error_count, ac_as_count


def _run_conversion() -> bool:
    """执行翻译转化：调用 process_orders_preserve_format 写入临时结果文件，
    更新会话统计并自动收录模板规则。返回是否成功。"""
    st.session_state.processing_done = False
    _cleanup(st.session_state.result_path)
    _ensure_callie_loaded()
    try:
        out_path, modified, skipped, error_count, ac_as_count = _convert_to_temp(st.session_state.orders_path)
        st.session_state.result_path = out_path
        st.session_state.stat_modified = modified
        st.session_state.stat_skipped = skipped
        st.session_state.stat_ac_as = ac_as_count
        st.session_state.exec_errors_list = (
            [{"来源": "订单执行", "错误类型": "规则执行异常",
              "错误原因": f"共有 {error_count} 行处理异常，详见终端日志"}]
            if error_count > 0 else []
        )
        st.session_state.processing_done = True
        # 自动收录模板规则
        if modified > 0 and st.session_state.template_rules:
            for sku, rules in st.session_state.template_rules.items():
                for r in rules:
                    _record_template(sku, r.original)
        _cleanup(st.session_state.get("_prev_result_path"))
        return True
    except Exception as e:
        st.error(f"❌ {e}")
        return False


# ╔════════════════════════════════════════════════════════════════╗
# ║  数据库模块：template_database.json                              ║
# ╚════════════════════════════════════════════════════════════════╝

import json
from datetime import date as _date

DATABASE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "template_database.json")
CALLIE_SKU_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "callie_sku_configs")


def _load_db() -> dict:
    """加载模板数据库并自动清理过期条目。返回 {SKU: [{template, source, added_at, use_count}, ...]}。"""
    db: dict = {}
    if os.path.exists(DATABASE_FILE):
        try:
            with open(DATABASE_FILE, "r", encoding="utf-8") as f:
                db = json.load(f)
        except Exception:
            return {}

    # 规则2：自动清理 added_at 超过 150 天的模板
    today = _date.today()
    cleaned = False
    for sku in list(db.keys()):
        entries = db[sku]
        kept = []
        for entry in entries:
            added_str = entry.get("added_at", "2000-01-01")
            try:
                added_date = _date.fromisoformat(added_str)
                if (today - added_date).days > 150:
                    logging.getLogger(__name__).info(
                        "已清理模板 %s…（SKU: %s，添加于 %s）",
                        entry.get("template", "")[:50], sku, added_str,
                    )
                    cleaned = True
                    continue
            except (ValueError, TypeError):
                pass
            kept.append(entry)
        if kept:
            db[sku] = kept
        else:
            del db[sku]

    if cleaned:
        _save_db(db)
    return db


def _save_db(db: dict) -> None:
    """保存模板数据库到 JSON 文件。"""
    with open(DATABASE_FILE, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, indent=2)


def _record_template(sku: str, template_text: str) -> None:
    """
    自动收录模板规则。

    规则1：单SKU最多10条，满时删除 added_at 最早的。
    规则3：数据库总条目 >= 5000 时停止收录。
    """
    if not sku or not template_text:
        return

    template_text = _normalize_template(template_text)

    db = _load_db()

    # 规则3：数据库总条目上限 5000
    total = sum(len(v) for v in db.values())
    if total >= 5000:
        logging.getLogger(__name__).warning("数据库已达上限（5000条），停止收录新模板")
        return

    if sku not in db:
        db[sku] = []

    entries = db[sku]
    today = str(_date.today())

    # 已存在 → use_count + 1
    for entry in entries:
        if entry["template"] == template_text:
            entry["use_count"] = entry.get("use_count", 0) + 1
            _save_db(db)
            return

    # 不存在 → 添加
    new_entry = {
        "template": template_text,
        "source": "自动收录",
        "added_at": today,
        "use_count": 1,
    }

    if len(entries) < 10:
        entries.append(new_entry)
    else:
        # 规则1：满10条时删除 added_at 最早的那条
        oldest_idx = min(
            range(len(entries)),
            key=lambda i: entries[i].get("added_at", "2000-01-01"),
        )
        removed = entries[oldest_idx]
        logging.getLogger(__name__).info(
            "已替换模板 %s…（SKU: %s，添加于 %s）",
            removed.get("template", "")[:50], sku, removed.get("added_at", ""),
        )
        entries[oldest_idx] = new_entry

    _save_db(db)


def _get_db_stats() -> tuple:
    """返回 (total_skus, total_templates)。"""
    db = _load_db()
    total_templates = sum(len(v) for v in db.values())
    return len(db), total_templates


def _normalize_template(template: str) -> str:
    """规范化模板文本：去除首尾空白，末尾补分号。"""
    template = template.strip()
    if template and not template.endswith(";"):
        template += ";"
    return template


def _format_template_multiline(tmpl: str) -> str:
    """将模板按 ; 分隔后，用换行连接以便多行显示，每条规则末尾保留分号。"""
    parts = [p.strip() + ";" for p in tmpl.split(";") if p.strip()]
    return "\n".join(parts)


# ╔════════════════════════════════════════════════════════════════╗
# ║  辅助：模板解析 → 调用 translator.load_template                  ║
# ╚════════════════════════════════════════════════════════════════╝

def _load_template_with_errors(file_path: str):
    """
    调用 translator.load_template 获取解析后的模板，
    同时读取原始数据生成错误报告。
    """
    # 核心解析（来自 translator.py）
    parsed = load_template(file_path)

    # 读取原始数据生成错误报告
    df = pd.read_excel(file_path, dtype=str)
    sku_col = _find_column(df, ["SKU", "sku", "商品编码"])
    rule_col = _find_column(df, ["翻译模板", "规则", "rule", "template"])
    if sku_col is None and len(df.columns) >= 2:
        sku_col = df.columns[1]
    if rule_col is None and len(df.columns) >= 3:
        rule_col = df.columns[2]
    if sku_col is None or rule_col is None:
        return parsed, df, [], sku_col, rule_col

    errors = []
    for idx, row in df.iterrows():
        sku = str(row[sku_col]).strip() if pd.notna(row[sku_col]) else ""
        rule_str = str(row[rule_col]).strip() if pd.notna(row[rule_col]) else ""
        excel_row = idx + 2
        if sku and rule_str:
            # 如果 SKU 不在解析结果中（或为空列表），说明解析失败
            rules = parsed.get(sku)
            if rules is None or len(rules) == 0:
                # 再次调用 parse_rule 获取具体错误信息
                try:
                    test = parse_rule(rule_str)
                    if not test:
                        errors.append({
                            "来源": "模板解析", "SKU": sku, "Excel行号": excel_row,
                            "问题规则": rule_str, "错误类型": "规则解析失败",
                            "错误原因": "规则解析后无有效结果",
                        })
                except Exception as e:
                    errors.append({
                        "来源": "模板解析", "SKU": sku, "Excel行号": excel_row,
                        "问题规则": rule_str, "错误类型": "规则解析异常",
                        "错误原因": str(e),
                    })
    return parsed, df, errors, sku_col, rule_col


def _parse_template_from_df(df: pd.DataFrame):
    """
    直接从 DataFrame 解析模板规则（跳过 Excel 文件 I/O 往返）。
    用于 CSV 粘贴数据，避免先保存为 Excel 再重新读取。
    """
    sku_col = _find_column(df, ["SKU", "sku", "商品编码"])
    rule_col = _find_column(df, ["翻译模板", "规则", "rule", "template"])
    if sku_col is None and len(df.columns) >= 2:
        sku_col = df.columns[1]
    if rule_col is None and len(df.columns) >= 3:
        rule_col = df.columns[2]
    if sku_col is None or rule_col is None:
        raise ValueError(
            f"无法识别模板中的SKU或规则列。可用列: {list(df.columns)}"
        )

    errors = []
    parsed = {}
    for idx, row in df.iterrows():
        sku = str(row[sku_col]).strip() if pd.notna(row[sku_col]) else ""
        rule_str = str(row[rule_col]).strip() if pd.notna(row[rule_col]) else ""
        excel_row = idx + 2
        if sku and rule_str:
            rules = parse_rule(rule_str)
            if rules:
                parsed[sku] = rules
            else:
                errors.append({
                    "来源": "模板解析", "SKU": sku, "Excel行号": excel_row,
                    "问题规则": rule_str, "错误类型": "规则解析失败",
                    "错误原因": "规则解析后无有效结果",
                })
        elif sku:
            parsed[sku] = []

    return parsed, df, errors, sku_col, rule_col


# ╔════════════════════════════════════════════════════════════════╗
# ║              横向四列布局                                      ║
# ╚════════════════════════════════════════════════════════════════╝

# ╔════════════════════════════════════════════════════════════════╗
# ║              横向四列布局（云端版：手动上传 → 转化 → 下载）      ║
# ╚════════════════════════════════════════════════════════════════╝

col1, col2, col3, col4 = st.columns(4, gap="small")

# ── 列1：上传模板 ────────────────────────────────────────────────
with col1, _step_card("步骤1：上传模板"):
    template_file = st.file_uploader(
        "拖拽或点击上传模板 Excel（支持 .xlsx / .xls）",
        type=["xlsx", "xls"], key="template_uploader", label_visibility="collapsed",
    )
    if template_file is not None:
        if _is_shortcut(template_file):
            st.error("❌ 上传的是 Windows 快捷方式（.lnk），不是真实 Excel。请右键快捷方式→「打开文件所在位置」，选真实的 .xlsx/.xls 上传。")
            st.session_state.template_saved = False
        else:
            _cleanup(st.session_state.result_path)
            tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            tmp.write(template_file.getbuffer())
            tmp.close()
            real_path = tmp.name
            try:
                parsed, df, errors, _, _ = _load_template_with_errors(real_path)
                st.session_state.processing_done = False
                st.session_state.template_rules = parsed
                st.session_state.template_df = df
                st.session_state.template_name = template_file.name
                st.session_state.template_saved = True
                st.session_state.template_path = real_path
                st.session_state.parse_errors_list = errors
                st.session_state.csv_template_text = ""
                _ensure_callie_loaded()
            except Exception as e:
                st.error(f"❌ {e}")
                st.session_state.template_saved = False
                _cleanup(real_path)

    if st.session_state.template_saved:
        st.markdown(
            f'<span class="saved-badge">✅ {st.session_state.template_name}'
            f'（{len(st.session_state.template_df)} 行）</span>',
            unsafe_allow_html=True,
        )
        if st.button("🗑️ 清除", key="clear_template", use_container_width=True):
            for k in ["template_df", "template_name", "template_saved", "template_rules",
                "parse_errors_list", "csv_template_text", "processing_done"]:
                st.session_state[k] = None if k.endswith("_df") or k.endswith("_rules") else (
                    [] if k == "parse_errors_list" else "" if k == "template_name" or k == "csv_template_text" else False
                )
            st.rerun()

# ── 列2：上传订单 ────────────────────────────────────────────────
with col2, _step_card("步骤2：上传订单"):
    orders_file = st.file_uploader(
        "拖拽或点击上传订单 Excel",
        type=["xlsx", "xls"], key="orders_uploader", label_visibility="collapsed",
    )
    if orders_file is not None:
        _cleanup(st.session_state.orders_path)
        _cleanup(st.session_state.result_path)
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.write(orders_file.getbuffer())
        tmp.close()
        try:
            df = pd.read_excel(tmp.name, dtype=str)
            st.session_state.orders_df = df
            st.session_state.orders_name = orders_file.name
            st.session_state.orders_path = tmp.name
            st.session_state.orders_saved = True
            st.session_state.csv_orders_text = ""
            st.session_state.processing_done = False
            st.session_state.auto_export_msg = None
        except Exception as e:
            st.error(f"❌ {e}")
            st.session_state.orders_saved = False
            _cleanup(tmp.name)

    if st.session_state.orders_saved:
        st.markdown(
            f'<span class="saved-badge">✅ {st.session_state.orders_name}'
            f'（{len(st.session_state.orders_df)} 行）</span>',
            unsafe_allow_html=True,
        )
        if st.button("🗑️ 清除", key="clear_orders", use_container_width=True):
            _cleanup(st.session_state.orders_path)
            _cleanup(st.session_state.result_path)
            for k in ["orders_df", "orders_name", "orders_saved", "orders_wb", "orders_path",
                "result_path", "csv_orders_text", "processing_done", "exec_errors_list"]:
                st.session_state[k] = None if k.endswith("_df") or k.endswith("_wb") or k.endswith("_path") else (
                    [] if k == "exec_errors_list" else "" if k == "orders_name" or k == "csv_orders_text" else False
                )
            st.session_state.auto_export_msg = None
            st.rerun()

# ── 列3：导出表格 ────────────────────────────────────────────────
with col3, _step_card("步骤3：导出表格"):
    both_ready = (
        st.session_state.template_saved and st.session_state.orders_saved
        and st.session_state.template_rules is not None
        and st.session_state.orders_path is not None
    )

    if not both_ready:
        st.info("👆 请先完成步骤1和2")
    else:
        if st.button("🔍 开始转化", type="primary", use_container_width=True, key="btn_convert"):
            _run_conversion()

        if st.session_state.processing_done and st.session_state.result_path:
            st.divider()
            sm1, sm2 = st.columns(2)
            with sm1:
                st.markdown(f'<div class="stat-box"><div class="number">{st.session_state.stat_modified}</div><div class="label">✅ 已修改</div></div>', unsafe_allow_html=True)
            with sm2:
                st.markdown(f'<div class="stat-box"><div class="number">{st.session_state.stat_skipped}</div><div class="label">⏭️ 已跳过</div></div>', unsafe_allow_html=True)
            sm3, sm4 = st.columns(2)
            with sm3:
                ec = len(st.session_state.exec_errors_list)
                c = "#dc2626" if ec > 0 else "#4f46e5"
                st.markdown(f'<div class="stat-box"><div class="number" style="color:{c}">{ec}</div><div class="label">❌ 错误</div></div>', unsafe_allow_html=True)
            with sm4:
                st.markdown(f'<div class="stat-box"><div class="number">{st.session_state.stat_modified + st.session_state.stat_skipped}</div><div class="label">📦 总计</div></div>', unsafe_allow_html=True)
            ac = st.session_state.stat_ac_as
            ac_color = "#4f46e5" if ac > 0 else "#dc2626"
            st.markdown(f'<div class="stat-box"><div class="number" style="color:{ac_color}">{ac}</div><div class="label">🔧 AC/AS 填充行数</div></div>', unsafe_allow_html=True)
            if ac == 0:
                st.warning("⚠️ AC/AS 自动填充数为 0！请确认：① 订单文件有 K 列(加急)；② 订单行有 SKU")
            ca = st.session_state.get("callie_section_count", 0)
            ca_color = "#4f46e5" if ca > 0 else "#9ca3af"
            st.markdown(f'<div class="stat-box"><div class="number" style="color:{ca_color}">{ca}</div><div class="label">🎨 callie定制项(独立版块)</div></div>', unsafe_allow_html=True)

            if st.session_state.result_path and os.path.exists(st.session_state.result_path):
                with open(st.session_state.result_path, "rb") as f:
                    result_bytes = f.read()
                st.download_button("📥 导出结果", data=result_bytes,
                    file_name=f"{os.path.splitext(st.session_state.orders_name)[0]}-副本.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key="download_result")

            if st.button("🔄 重新开始", use_container_width=True, key="reset_all"):
                _cleanup(st.session_state.get("orders_path"))
                _cleanup(st.session_state.get("result_path"))
                for key in list(st.session_state.keys()):
                    del st.session_state[key]
                st.rerun()

# ── 列4：错误报告 ────────────────────────────────────────────────
with col4, _step_card("步骤4：错误报告"):
    pe = st.session_state.parse_errors_list
    ee = st.session_state.exec_errors_list
    if not pe and not ee:
        if st.session_state.processing_done:
            st.success("🎉 无错误")
        else:
            st.info("等待执行翻译")
    else:
        all_errs = pe + ee
        e_df = pd.DataFrame(all_errs)
        st.warning(f"**{len(e_df)} 条错误**（解析{len(pe)} + 执行{len(ee)}）")
        with st.expander("📋 展开详情", expanded=False):
            st.dataframe(e_df, use_container_width=True, hide_index=True,
                column_config={
                    "来源": st.column_config.TextColumn("来源", width="small"),
                    "SKU": st.column_config.TextColumn("SKU", width="medium"),
                    "Excel行号": st.column_config.NumberColumn("行号", width="small"),
                    "问题规则": st.column_config.TextColumn("问题规则", width="large"),
                    "错误类型": st.column_config.TextColumn("错误类型", width="medium"),
                    "错误原因": st.column_config.TextColumn("错误原因", width="large"),
                })

# ── A 系统规则草稿生成（独立面板，不依赖步骤一）──
with st.expander("规则草稿生成"):
    import translator as tr
    import streamlit.components.v1 as components
    import json

    _c1, _c2, _c3 = st.columns(3)

    # 统一标题行高度，保证三个文本框顶部对齐
    _TITLE_H = 34

    with _c1:
        st.markdown(
            f"<div style='height:{_TITLE_H}px;display:flex;align-items:center;'>"
            f"<b>📝 翻译前定制项</b></div>",
            unsafe_allow_html=True,
        )
        _b = st.text_area("翻译前定制项", value=st.session_state.get("gen_a_before", ""),
                          key="gen_a_before", label_visibility="collapsed",
                          placeholder="（J列样例）")

    with _c2:
        st.markdown(
            f"<div style='height:{_TITLE_H}px;display:flex;align-items:center;'>"
            f"<b>📋 翻译后定制项</b></div>",
            unsafe_allow_html=True,
        )
        _a = st.text_area("翻译后定制项", value=st.session_state.get("gen_a_after", ""),
                          key="gen_a_after", label_visibility="collapsed",
                          placeholder="（AZ列样例）")

    with _c3:
        _fb = json.dumps(st.session_state.get("generator_rules", ""))
        # 标题 + 悬浮复制按钮；按钮事件在下方 components.html 中绑定，不占额外高度
        st.markdown(
            f"<div style='height:{_TITLE_H}px;display:flex;align-items:center;position:relative;'>"
            f"<b>⚙️ 规则草稿生成</b>"
            f"<button id='copyDraftBtn' type='button' title='复制草稿' "
            f"style='position:absolute;right:0;top:50%;transform:translateY(-50%);border:none;background:transparent;font-size:18px;cursor:pointer;padding:2px 6px;border-radius:6px;' "
            f"onmouseover='this.style.background=\"#e2e8f0\"' "
            f"onmouseout='this.style.background=\"transparent\"'>📋</button>"
            f"</div>",
            unsafe_allow_html=True,
        )
        _draft_val = st.text_area("翻译模板草稿", value=st.session_state.get("generator_rules", ""),
                                  key="gen_a_rules", label_visibility="collapsed",
                                  placeholder="（生成后显示规则草稿）")

    # 生成 / 校验按钮并排
    _btn1, _btn2 = st.columns(2)
    with _btn1:
        if st.button("📝 生成草稿", key="gen_a_btn", use_container_width=True):
            _bb = st.session_state.gen_a_before
            _aa = st.session_state.gen_a_after
            if not _bb.strip() or not _aa.strip():
                st.warning("请先填写「翻译前」和「翻译后」定制项")
            else:
                try:
                    _draft = tr.generate_a_rule_draft(_bb, _aa)
                    st.session_state.generator_rules = _draft
                    st.rerun()
                except Exception as _e:
                    st.error(f"生成失败：{_e}")
    with _btn2:
        if st.button("✅ 校验", key="gen_a_test", use_container_width=True):
            _draft_val = st.session_state.get("gen_a_rules", "")
            if not _draft_val.strip():
                st.warning("草稿为空，请先点「生成草稿」")
            else:
                try:
                    _rules = tr.parse_rule(tr.strip_rule_comments(_draft_val))
                    _sku, _result = tr.apply_rules(_rules, "SKU", st.session_state.gen_a_before)
                    def _norm(s):
                        import re
                        return [re.sub(r"\s*:\s*", ":", l).strip()
                                for l in str(s).splitlines() if l.strip()]
                    if _norm(_result) == _norm(st.session_state.gen_a_after):
                        st.success("✅ 草稿作用于「翻译前」后，结果与「翻译后」完全一致")
                    else:
                        st.warning("⚠️ 有差异：")
                        st.code("—— 草稿生成结果 ——\n" + "\n".join(_norm(_result)))
                        st.code("—— 期望（翻译后）——\n" + "\n".join(_norm(st.session_state.gen_a_after)))
                except Exception as _e:
                    st.error(f"编译/校验失败：{_e}")

    # 三个输入框高度自适应（随内容增长），并绑定复制按钮事件
    _auto_resize_html = (
        "<script>"
        "var _fb=" + _fb + ";"
        "function _ar(ta){ta.style.height='auto';ta.style.height=Math.max(ta.scrollHeight,100)+'px';}"
        "function _bindCopy(){"
        "var b=window.parent.document.getElementById('copyDraftBtn');"
        "var ta=window.parent.document.querySelector('textarea[aria-label=\"翻译模板草稿\"]');"
        "if(b&&!b.dataset.bound){b.dataset.bound='1';"
        "b.addEventListener('click',function(){"
        "var txt=ta?ta.value:_fb;"
        "if(txt===undefined||txt===null)txt=_fb;"
        "navigator.clipboard.writeText(txt);"
        "b.textContent='✅';setTimeout(function(){b.textContent='📋';},1500);"
        "});}"
        "}"
        "function _initAR(){"
        "['翻译前定制项','翻译后定制项','翻译模板草稿'].forEach(function(l){"
        "var ta=window.parent.document.querySelector('textarea[aria-label=\"'+l+'\"]');"
        "if(ta&&!ta.dataset.ar){ta.dataset.ar='1';ta.addEventListener('input',function(){_ar(ta);});_ar(ta);}"
        "});"
        "_bindCopy();"
        "}"
        "_initAR();"
        "var _obs=new MutationObserver(_initAR);"
        "_obs.observe(window.parent.document.body,{childList:true,subtree:true});"
        "</script>"
    )
    components.html(_auto_resize_html, height=0)


# ════════════════════════════════════════════════════════════════════════════
# 独立功能：Callie定制项生成（上传订单 + 翻译模板 → 生成 AY 列）
# 规则由翻译模板的 D/E/F 列驱动，经 build_callie_rule.py 编译后逐行生成 AY。
# 属性表已持久化到 callie_sku_configs/<SKU>/，无需每次上传。
# ════════════════════════════════════════════════════════════════════════════
st.divider()

# B 系统 header + 右上角指示标（规则说明 / 已入库属性表）
_b_left, _b_mid, _b_right = st.columns([5, 1, 1])
with _b_left:
    st.markdown("""
    <div class="main-header" style="background: linear-gradient(135deg, #0ea5e9 0%, #6366f1 100%);">
        <h1>🎨 callie定制项填充</h1>
    </div>
    """, unsafe_allow_html=True)
with _b_mid:
    st.markdown('<div style="height:18px"></div>', unsafe_allow_html=True)
    if st.button("📚 已入库属性表", key="btn_registered_skus", use_container_width=True):
        _show_registered_skus()
with _b_right:
    st.markdown('<div style="height:18px"></div>', unsafe_allow_html=True)
    if st.button("📖 规则说明", key="btn_rules_b", use_container_width=True):
        _show_rules_b()

def _derive_sku_from_filename(name):
    """从上传文件名推断 SKU（去掉扩展名）。用户把 attribute.xlsx 直接命名为 SKU 即可免手填。"""
    return os.path.splitext(os.path.basename(str(name)))[0].strip()

with st.expander("属性表入库"):
    attr_file = st.file_uploader("attribute.xlsx（文件名 = SKU 编号）", type=["xlsx"], key="new_attr")
    if st.button("📥 入库属性表", key="btn_register"):
        if not attr_file:
            st.warning("请上传 attribute.xlsx")
        elif _is_shortcut(attr_file):
            st.error("❌ 上传的是 Windows 快捷方式（.lnk），不是真实 Excel。请选真实 .xlsx 上传。")
        else:
            sku = _derive_sku_from_filename(attr_file.name)
            if not sku:
                st.warning("无法从文件名推断 SKU，请把文件命名为 SKU 编号后重新上传")
            else:
                tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
                tmp.write(attr_file.getbuffer()); tmp.close()
                try:
                    import build_callie_sku
                    attr = build_callie_sku.main(sku, tmp.name)
                    n_grp = len(attr.get("groups", [])) if attr else 0
                    n_opt = sum(1 for g in attr.get("groups", []) if g.get("opts")) if attr else 0
                    if n_opt == 0:
                        st.warning(f"✅ 已入库 {sku}（{n_grp} 组），但**没有任何选项列表(opts)**——发色路由等值映射可能报「固定ID模式」。请确认上传的是 callie 后台导出的完整属性表。")
                    else:
                        st.success(f"✅ 已入库 {sku}（{n_grp} 组 / 含选项 {n_opt} 组），B 系统导出时自动复用")
                except Exception as e:
                    st.error(f"入库失败：{e}")
                finally:
                    _cleanup(tmp.name)

# ── 规则草稿生成（独立上传模板，不依赖步骤一）──
with st.expander("规则草稿生成"):
    import build_callie_rule as bcr
    import json as _json

    # 独立上传翻译模板：本面板自取，不依赖步骤一
    _draft_upl = st.file_uploader(
        "拖拽或点击上传翻译模板 Excel（支持 .xlsx / .xls）",
        type=["xlsx", "xls"], key="draft_template", label_visibility="collapsed",
    )
    if _draft_upl is not None:
        if _is_shortcut(_draft_upl):
            st.error("❌ 上传的是 Windows 快捷方式（.lnk），不是真实 Excel。请选真实 .xlsx/.xls 上传。")
        else:
            _cleanup(st.session_state.get("draft_template_path"))
            _dn = _draft_upl.name
            _tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            _tmp.write(_draft_upl.getbuffer())
            _tmp.close()
            _rp = _tmp.name
            st.session_state.draft_template_path = _rp
            st.session_state.draft_template_name = _dn
            st.session_state.draft_template_saved = True
            # 同步给步骤三导出用，省得再传一次
            st.session_state.callie_template_path = _rp
            st.session_state.callie_template_name = _dn
            st.session_state.callie_template_saved = True

    if not st.session_state.get("draft_template_saved"):
        if st.session_state.get("callie_template_saved") and st.session_state.get("callie_template_path"):
            st.session_state.draft_template_path = st.session_state.callie_template_path
            st.session_state.draft_template_saved = True
            st.info("已复用步骤一上传的翻译模板；如需换模板，可在此重新上传。")
        else:
            st.info("👆 请上传翻译模板，系统才能读取各 SKU 的 E（J 示例）/ F（AZ 示例）列")
    else:
        _tpl_path = st.session_state.draft_template_path
        try:
            _tmpl = bcr.read_template(_tpl_path)
        except Exception as _e:
            st.error(f"读取模板失败：{_e}")
            _tmpl = {}
        if not _tmpl:
            st.warning("模板里没有任何 SKU 的 D/E/F 内容")
        else:
            _sku = st.selectbox("选择 SKU", list(_tmpl.keys()), key="gen_sku")
            _cols = _tmpl[_sku]
            st.caption("当前 标准化规则（阶段 0）：")
            st.code(_cols.get("STD") or "(空)", language="text")
            st.caption("当前 callie定制项翻译规则：")
            st.code(_cols["D"] or "(空)", language="text")

            def _load_attr_for(sku):
                _lib = os.path.join(CALLIE_SKU_DIR, sku, "attribute_config.json")
                if os.path.isfile(_lib):
                    with open(_lib, encoding="utf-8") as _fh:
                        return _json.load(_fh)
                return None

            def _auto_h(text, min_h=120, line_px=21, pad=20):
                # 按内容行数动态撑开高度，避免固定高度截断内容
                return max(min_h, len(str(text).splitlines()) * line_px + pad)

            _e = st.text_area("临时列(需删除) 示例（J）", value=_cols["E"], key="gen_e", height=_auto_h(_cols["E"]))
            _f = st.text_area("callie定制项 示例（AZ）", value=_cols["F"], key="gen_f", height=_auto_h(_cols["F"], min_h=160))

            if st.button("🤖 生成草稿", key="gen_btn", use_container_width=True):
                try:
                    _attr = _load_attr_for(_sku)
                    _draft = bcr.generate_rule_draft(_e, _f, _attr)
                    st.session_state.gen_draft_ta = _draft
                    st.rerun()
                except Exception as _e:
                    st.error(f"生成失败：{_e}")

            _draft_val = st.text_area(
                "callie定制项翻译规则草稿（可编辑，改完点「应用到该 SKU」）",
                value=st.session_state.get("gen_draft_ta", ""),
                key="gen_draft_ta", height=_auto_h(st.session_state.get("gen_draft_ta", ""), min_h=200),
            )
            _std_val = st.text_area(
                "标准化规则（阶段 0，可编辑，每条 rule|原始行|目标行…；应用到该 SKU 时一并写回）",
                value=_cols.get("STD", ""),
                key="gen_std_ta", height=_auto_h(_cols.get("STD", ""), min_h=120),
            )

            _c1, _c2 = st.columns(2)
            with _c1:
                if st.button("▶ 对示例跑一遍（期望 AZ vs 生成 AZ）", key="gen_test", use_container_width=True):
                    if not _draft_val.strip():
                        st.warning("草稿为空，请先点「🤖 生成草稿」")
                    else:
                        try:
                            _attr = _load_attr_for(_sku)
                            if _attr is None:
                                st.warning("⚠️ 该 SKU 未入库属性表，无法编译校验；草稿仍可直接应用")
                            _rule = bcr.compile_rule(_sku, _draft_val, _e, _f, _attr)
                            _ok, _diffs = bcr.validate_rule(_rule, _e, _f, _attr)
                            if _ok:
                                st.success("✅ 草稿生成的 AZ 与示例 F 完全一致")
                            else:
                                st.warning("⚠️ 草稿生成的 AZ 与示例 F 有差异（修改罗盘）：")
                                for _d in _diffs:
                                    st.code(_d)
                        except Exception as _e:
                            st.error(f"编译/校验失败：{_e}")
            with _c2:
                if st.button("💾 应用到该 SKU", key="gen_apply", use_container_width=True):
                    if not _draft_val.strip() and not _std_val.strip():
                        st.warning("草稿与标准化规则均为空")
                    else:
                        try:
                            import openpyxl
                            _wb = openpyxl.load_workbook(_tpl_path)
                            _ws = _wb.active
                            # 按表头名定位 SKU / 映射规则 / 标准化规则 列，避免写错列
                            _dcols = bcr._detect_columns(_tpl_path)
                            _sku_c, _map_c, _std_c = _dcols["sku"], _dcols["mapping"], _dcols["std"]
                            for _row in _ws.iter_rows():
                                if _sku_c < len(_row) and str(_row[_sku_c].value or "").strip() == _sku:
                                    if _draft_val.strip() and _map_c is not None and _map_c < len(_row):
                                        _row[_map_c].value = _draft_val
                                    if _std_val.strip() and _std_c is not None and _std_c < len(_row):
                                        _row[_std_c].value = _std_val
                                    break
                            _wb.save(_tpl_path)
                            st.session_state.callie_result_path = None
                            st.success(f"✅ 已写入 {_sku} 的 callie定制项翻译规则列" +
                                       (" 与 标准化规则列" if _std_val.strip() else "") +
                                       "，下次点「🚀 导出表格」即用此规则")
                            st.rerun()
                        except Exception as _e:
                            st.error(f"写回失败（模板需为 .xlsx 格式）：{_e}")

def _save_upload(uploaded):
    """把上传的文件存成临时文件，返回路径（openpyxl/pandas 需要真实路径）。"""
    suffix = os.path.splitext(uploaded.name)[1] or ".xlsx"
    tmp = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
    tmp.write(uploaded.getbuffer())
    tmp.close()
    return tmp.name


def _is_shortcut(uploaded):
    """判断上传的究竟是真实 Excel，还是 Windows 快捷方式（.lnk）。

    .lnk 文件被拖进上传框时，Streamlit 只看到 1KB 左右的快捷方式字节，
    openpyxl/pandas 无法解析，会报「文件不是有效 xlsx」之类错误。
    通过文件名后缀 + 文件头魔数双重判定，给出清晰提示。
    """
    name = (uploaded.name or "").lower()
    if name.endswith(".lnk"):
        return True
    # 文件头魔数：.lnk 固定以字节 4C 00 00 00 01 14 02 00 开头
    head = bytes(uploaded.getbuffer()[:8])
    if head[:4] == b"\x4c\x00\x00\x00" and head[4:8] == b"\x01\x14\x02\x00":
        return True
    return False


def _append_missing_columns(dst_path, result_df, missing_cols):
    """原订单缺失的同步列（AW/AX/AY 等）追加到末尾，保证导出文件包含并填值。

    仅在轻量更新之后调用；对「列缺失」的老订单兜底，避免数据静默丢失。
    用 openpyxl 追加（保留已轻量写入的样式/图片），效率低于 XML 直改但仅针对少量缺列。
    """
    import openpyxl
    wb = openpyxl.load_workbook(dst_path)
    ws = wb.active
    start_col = ws.max_column + 1
    n_rows = len(result_df)
    for offset, col in enumerate(missing_cols):
        c_idx = start_col + offset
        # 表头
        ws.cell(row=1, column=c_idx, value=col)
        # 数据行（result_df 与订单行一一对应）
        for i in range(n_rows):
            v = result_df.iloc[i][col]
            ws.cell(row=i + 2, column=c_idx, value="" if pd.isna(v) else str(v))
    wb.save(dst_path)


def _export_callie_preserve_format(src_path, dst_path, orig_df, result_df,
                                   sync_cols=("callie定制项", "参考callie站点", "calie商品ID", "callie商品版本")):
    """保留原订单格式（图片/列宽/样式/合并单元格）导出，只写有变化的列单元格。

    复用 translator 的 zipfile 直改 XML 方式（_update_xlsx_cells_lightweight），
    不经过 pandas 写回，避免 to_excel 丢失图片与格式。

    同步列（默认，列位置以原始订单 pet8.17.xls 为准）：
      callie定制项(AZ)            —— callie定制项翻译规则列生成
      参考callie站点(AW)          —— 恒为 callie
      calie商品ID(AX)            —— 翻译模板「calie商品ID和商品版本」列
      callie商品版本(AY)          —— 翻译模板「calie商品ID和商品版本」列

    鲁棒性：
      * 原订单已含这些列 → 轻量更新已有单元格（保留图片/格式）；
      * 原订单缺某列（老订单/不同模板导出）→ 该列追加到末尾（列名 + 每行值），保证数据不丢；
      * 全部列都不存在 → 退回 pandas 写（会丢失图片/格式）。
    """
    present = [c for c in sync_cols if c in orig_df.columns and c in result_df.columns]
    missing = [c for c in sync_cols if c in result_df.columns and c not in orig_df.columns]

    if not present and not missing:
        # 兜底：老订单无这些列 → 退回 pandas 写
        result_df.to_excel(dst_path, index=False)
        return

    updates = {}
    for col in present:
        ci = list(orig_df.columns).index(col)
        letter = _col_idx_to_letter(ci + 1)
        for i in range(len(orig_df)):
            old_v = orig_df.iloc[i][col]
            new_v = result_df.iloc[i][col]
            old_s = "" if pd.isna(old_v) else str(old_v)
            new_s = "" if pd.isna(new_v) else str(new_v)
            if old_s != new_s:
                # 0-based 数据行 → 1-based Excel 行（+1 表头 +1 索引）
                updates.setdefault(i + 2, {})[letter] = new_s

    if not updates and not missing:
        shutil.copy2(src_path, dst_path)
        return

    # 对 .xls 等旧二进制格式无法直改 XML，先转成临时 xlsx 再轻量写入
    #（.xls 本身不支持图片/现代格式，转 xlsx 仅损失原 .xls 的格式，数据保留）
    import zipfile
    src_is_xlsx = False
    try:
        with zipfile.ZipFile(src_path, "r"):
            src_is_xlsx = True
    except Exception:
        src_is_xlsx = False

    if src_is_xlsx:
        _update_xlsx_cells_lightweight(src_path, dst_path, updates)
    else:
        tmp_xlsx = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
        try:
            orig_df.to_excel(tmp_xlsx, index=False, engine="openpyxl")
            _update_xlsx_cells_lightweight(tmp_xlsx, dst_path, updates)
        finally:
            _cleanup(tmp_xlsx)

    # 原订单缺失的列（AW/AX/AY 等）→ 追加到末尾，确保导出文件里出现并填值
    if missing:
        _append_missing_columns(dst_path, result_df, missing)

# ── B 系统四列卡片布局（与 A 系统一致）────────────────────────────
bcol1, bcol2, bcol3, bcol4 = st.columns(4, gap="small")

# 列1：上传翻译模板（顺序与 A 系统一致：先模板，再订单）
with bcol1, _step_card("步骤1：上传翻译模板"):
    template_file = st.file_uploader(
        "拖拽或点击上传翻译模板 Excel（支持 .xlsx / .xls）",
        type=["xlsx", "xls"], key="callie_template", label_visibility="collapsed",
    )
    if template_file is not None:
        if _is_shortcut(template_file):
            st.error("❌ 上传的是 Windows 快捷方式（.lnk），不是真实 Excel。请右键快捷方式→「打开文件所在位置」，选真实的 .xlsx/.xls 上传。")
            st.session_state.callie_template_saved = False
        else:
            _cleanup(st.session_state.get("callie_template_path"))
            tpl_name = template_file.name
            tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            tmp.write(template_file.getbuffer())
            tmp.close()
            real_path = tmp.name
            try:
                st.session_state.callie_template_path = real_path
                st.session_state.callie_template_name = tpl_name
                st.session_state.callie_template_saved = True
                st.session_state.callie_result_path = None
            except Exception as e:
                st.error(f"❌ {e}")
                st.session_state.callie_template_saved = False
                _cleanup(real_path)

    if st.session_state.get("callie_template_saved"):
        st.markdown(
            f'<span class="saved-badge">✅ {st.session_state.callie_template_name}</span>',
            unsafe_allow_html=True,
        )
        if st.button("🗑️ 清除", key="clear_callie_template", use_container_width=True):
            _cleanup(st.session_state.get("callie_template_path"))
            _cleanup(st.session_state.get("callie_result_path"))
            for k in ["callie_template_path", "callie_template_name", "callie_template_saved",
                      "callie_result_path", "callie_errors", "callie_cfg_summary"]:
                if k.endswith("_path"):
                    st.session_state[k] = None
                elif k in ("callie_errors", "callie_cfg_summary"):
                    st.session_state[k] = []
                elif k.endswith("_saved"):
                    st.session_state[k] = False
                else:
                    st.session_state[k] = ""
            st.rerun()

# 列2：上传订单
with bcol2, _step_card("步骤2：上传订单"):
    order_file = st.file_uploader(
        "拖拽或点击上传订单 Excel（含 SKU 列和「临时列(需删除)」列）",
        type=["xlsx", "xls"], key="callie_order", label_visibility="collapsed",
    )
    if order_file is not None:
        if _is_shortcut(order_file):
            st.error("❌ 上传的是 Windows 快捷方式（.lnk），不是真实 Excel。请右键快捷方式→「打开文件所在位置」，选真实的 .xlsx/.xls 上传。")
            st.session_state.callie_order_saved = False
        else:
            _cleanup(st.session_state.get("callie_order_path"))
            _cleanup(st.session_state.get("callie_result_path"))
            tmp = tempfile.NamedTemporaryFile(suffix=os.path.splitext(order_file.name)[1] or ".xlsx", delete=False)
            tmp.write(order_file.getbuffer()); tmp.close()
            try:
                st.session_state.callie_order_df = pd.read_excel(tmp.name, dtype=str)
                st.session_state.callie_order_path = tmp.name
                st.session_state.callie_order_name = order_file.name
                st.session_state.callie_order_saved = True
                st.session_state.callie_result_path = None
            except Exception as e:
                st.error(f"❌ {e}")
                st.session_state.callie_order_saved = False
                _cleanup(tmp.name)

    if st.session_state.get("callie_order_saved"):
        st.markdown(
            f'<span class="saved-badge">✅ {st.session_state.callie_order_name}'
            f'（{len(st.session_state.callie_order_df)} 行）</span>',
            unsafe_allow_html=True,
        )
        if st.button("🗑️ 清除", key="clear_callie_order", use_container_width=True):
            _cleanup(st.session_state.get("callie_order_path"))
            _cleanup(st.session_state.get("callie_result_path"))
            for k in ["callie_order_df", "callie_order_path", "callie_order_name",
                      "callie_order_saved", "callie_result_path", "callie_errors",
                      "callie_cfg_summary"]:
                if k.endswith("_df") or k.endswith("_path"):
                    st.session_state[k] = None
                elif k in ("callie_errors", "callie_cfg_summary"):
                    st.session_state[k] = []
                elif k.endswith("_saved"):
                    st.session_state[k] = False
                else:
                    st.session_state[k] = ""
            st.rerun()

# 列3：导出表格
with bcol3, _step_card("步骤3：导出表格"):
    callie_both_ready = (
        st.session_state.get("callie_order_saved")
        and st.session_state.get("callie_template_saved")
        and st.session_state.get("callie_order_path")
        and st.session_state.get("callie_template_path")
    )

    if not callie_both_ready:
        st.info("👆 请先完成步骤1和2")
    else:
        if st.button("🚀 导出表格", type="primary", use_container_width=True, key="callie_export"):
            import build_callie_rule as bcr
            tmpl_path = order_path = None
            reused_from_lib = []
            try:
                with st.spinner("正在编译规则并生成 AZ 列..."):
                    tmpl_path = st.session_state.callie_template_path
                    order_path = st.session_state.callie_order_path

                    # 读翻译模板，确定需要哪些 SKU
                    template = bcr.read_template(tmpl_path)

                    # 翻译模板「calie商品ID和商品版本」列映射：{sku: {id, version}}（供填充 AW/AX/AY）
                    callie_product_map = bcr.build_callie_product_map(tmpl_path)

                    # 属性表来源：复用已入库的（callie_sku_configs/<SKU>/attribute_config.json）
                    attr_sources = {}
                    for sku in template:
                        lib = os.path.join(CALLIE_SKU_DIR, sku, "attribute_config.json")
                        if os.path.isfile(lib):
                            attr_sources[sku] = lib
                            reused_from_lib.append(sku)
                        # 否则不放 → compile_from_template 标记「缺少属性表」

                    # 编译各 SKU 规则
                    cfgs = bcr.compile_from_template(tmpl_path, attr_sources)

                    # 读订单
                    order_df = pd.read_excel(order_path, dtype=str)

                    # 逐行生成 AZ(callie定制项) + 填充 AW/AX(callie 商品信息)
                    result_df, errors = bcr.process_orders(order_df, cfgs,
                                                          callie_product_map=callie_product_map)

                    # 导出临时 xlsx（保留原订单格式：图片/列宽/样式/合并单元格）
                    out_path = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
                    _export_callie_preserve_format(order_path, out_path, order_df, result_df)

                st.session_state.callie_result_path = out_path
                st.session_state.callie_errors = errors
                st.session_state.callie_cfg_summary = {s: (c.get("error") or "✅") for s, c in cfgs.items()}
                if reused_from_lib:
                    st.session_state.callie_reused_skus = reused_from_lib
                else:
                    st.session_state.callie_reused_skus = []
                st.session_state.callie_product_map_len = len(callie_product_map)
            except Exception as e:
                st.error(f"❌ 处理失败：{e}")
            finally:
                # 注意：不要清理 session_state 中的模板/订单路径，它们后续还要用
                pass

        # 结果展示：统计 + 下载
        if st.session_state.get("callie_result_path"):
            errors = st.session_state.get("callie_errors") or []
            shown_errors = [e for e in errors if e.get("状态") != "未配置"]
            fail = [e for e in shown_errors if e["状态"] in ("生成失败", "异常", "配置错误")]
            summary = st.session_state.get("callie_cfg_summary") or {}

            # 规则编译失败必须显眼报出：否则 AZ 整列空却看不到原因（2026-08-21 踩坑）
            bad_cfg = {s: v for s, v in summary.items() if v != "✅"}
            if bad_cfg:
                st.error(
                    "⚠️ 以下 SKU 规则编译失败，**AZ（callie定制项）不会生成**：\n\n"
                    + "\n".join(f"- `{s}`：{v}" for s, v in bad_cfg.items())
                )

            # 统计：成功 SKU 数 / 失败行数
            ok_skus = sum(1 for v in summary.values() if v == "✅")
            total_skus = len(summary)
            sm1, sm2 = st.columns(2)
            with sm1:
                st.markdown(f'<div class="stat-box"><div class="number">{ok_skus}/{total_skus}</div><div class="label">✅ SKU 规则编译</div></div>', unsafe_allow_html=True)
            with sm2:
                c = "#dc2626" if fail else "#4f46e5"
                st.markdown(f'<div class="stat-box"><div class="number" style="color:{c}">{len(fail)}</div><div class="label">❌ 错误行</div></div>', unsafe_allow_html=True)

            if st.session_state.get("callie_reused_skus"):
                st.caption("♻️ 已复用属性表：" + "、".join(st.session_state.callie_reused_skus))
            if st.session_state.get("callie_product_map_len"):
                st.caption(f"📎 已从翻译模板「calie商品ID和商品版本」列匹配 {st.session_state.callie_product_map_len} 个 SKU 的商品信息")

            base_name = os.path.splitext(st.session_state.get("callie_order_name", "订单"))[0]
            with open(st.session_state.callie_result_path, "rb") as f:
                st.download_button("📥 下载订单副本（含 callie定制项列）", data=f.read(),
                                   file_name=f"{base_name}+副本2.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True, key="callie_dl")

            if st.button("🔄 重新开始", use_container_width=True, key="reset_callie"):
                _cleanup(st.session_state.get("callie_order_path"))
                _cleanup(st.session_state.get("callie_template_path"))
                _cleanup(st.session_state.get("callie_result_path"))
                for k in ["callie_order_df", "callie_order_path", "callie_order_name", "callie_order_saved",
                          "callie_template_path", "callie_template_name", "callie_template_saved",
                          "callie_result_path", "callie_errors", "callie_cfg_summary",
                          "callie_reused_skus", "callie_product_map_len"]:
                    if k.endswith("_df") or k.endswith("_path"):
                        st.session_state[k] = None
                    elif k in ("callie_errors", "callie_cfg_summary", "callie_reused_skus"):
                        st.session_state[k] = []
                    elif k.endswith("_saved"):
                        st.session_state[k] = False
                    else:
                        st.session_state[k] = ""
                st.rerun()

# 列4：错误报告
with bcol4, _step_card("步骤4：错误报告"):

    if not st.session_state.get("callie_result_path"):
        st.info("等待执行导出")
    else:
        errors = st.session_state.get("callie_errors") or []
        shown_errors = [e for e in errors if e.get("状态") != "未配置"]
        if not shown_errors:
            st.success("🎉 无错误")
        else:
            fail = [e for e in shown_errors if e["状态"] in ("生成失败", "异常", "配置错误")]
            st.warning(f"**{len(shown_errors)} 条提示**（严重 {len(fail)} 条）")
            with st.expander("📋 展开详情", expanded=False):
                st.dataframe(pd.DataFrame(shown_errors), use_container_width=True, hide_index=True,
                             column_config={
                                 "行号": st.column_config.NumberColumn("行号", width="small"),
                                 "SKU": st.column_config.TextColumn("SKU", width="medium"),
                                 "状态": st.column_config.TextColumn("状态", width="small"),
                                 "说明": st.column_config.TextColumn("说明", width="large"),
                             })

