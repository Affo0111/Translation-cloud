# -*- coding: utf-8 -*-
"""
CAPS250123 配置/规则生成器。

数据已对齐（截图 2026-08-26）：
  Title 长度 2/3/4/5/6 对应：
    NameColor gid = 162, 176, 190, 204, 218
    Title     gid = 163, 177, 191, 205, 219
    Name 基址  = 150, 164, 178, 192, 206（每长度 12 个连续 gid）

Name 的 12 个颜色选项（顺序决定基址偏移）：
  Black, White, Dark Purple, Dark Green, Blue, Brown,
  Light Pink, Pink, Dark Red, Gray, Light Blue, Green

逻辑：
  1) Title 长度决定 NameColor / Title 的 gid；
  2) Title 长度 + Name Color 颜色共同决定 Name 的 gid；
  3) Name 的 AZ 标签使用颜色名小写（如 dark purple:），值取 J 列「Name」字段（客人输入文本）。
"""
import json, os, shutil

HERE = os.path.dirname(os.path.abspath(__file__))
SKU = "CAPS250123"
OUTDIR = os.path.join(HERE, "callie_sku_configs", SKU)
os.makedirs(OUTDIR, exist_ok=True)

# ── 5 长度 gid 表 ──
NAMECOLOR = {2: 162, 3: 176, 4: 190, 5: 204, 6: 218}
NAME_BASE = {2: 150, 3: 164, 4: 178, 5: 192, 6: 206}
TITLE     = {2: 163, 3: 177, 4: 191, 5: 205, 6: 219}

# 12 个 name 颜色选项（顺序必须与 callie 选项→gid 顺序一致）
NAME_OPTS = [
    "Black", "White", "Dark Purple", "Dark Green", "Blue", "Brown",
    "Light Pink", "Pink", "Dark Red", "Gray", "Light Blue", "Green",
]
# AZ 列 label 使用小写颜色名（与模板 F 列示例 `180|dark purple:|Marcie` 一致）
LABEL_FOR = {
    "Black": "black", "White": "white", "Dark Purple": "dark purple",
    "Dark Green": "dark green", "Blue": "blue", "Brown": "brown",
    "Light Pink": "light pink", "Pink": "pink", "Dark Red": "dark red",
    "Gray": "gray", "Light Blue": "light blue", "Green": "green",
}

# 静态字段 gid（来自模板 F 列 len4 示例；这些不随 Title 长度变化）
STATIC = {
    "Gender": 100, "Badge Reel Color": 221, "Facial Expression": 122,
    "Skin Tone": 124, "Eyes Color": 123, "Glasses": 101, "Face Mask": 117,
    "Hair Color": 107, "Hairstyle": 116, "Undershirt": 119, "Pants": 120,
    "Clothes": 118, "Shoes": 121, "Hand-held Item": 103, "Drinks": 106,
}

# ── 1) attribute_config.json：只读，绝不覆盖 ──
# 属性表唯一来源 = callie 后台导出 xlsx 经 B 系统「属性表入库」(build_callie_sku.main) 编译生成。
# 本脚本只负责生成【规则文本】，不再硬编码重写属性表，避免覆盖已入库的完整版。
#（2026-08-26 事故：此处曾硬编码写残缺 gid 表，覆盖了用户已入库的 123-group 完整属性表，
#   导致发色→发型 gid 路由报「固定ID模式」。现已改为只读。）
attr_path = os.path.join(OUTDIR, "attribute_config.json")
if not os.path.exists(attr_path):
    raise SystemExit(
        "❌ 缺少 attribute_config.json。请先在 B 系统「属性表入库」上传 callie 导出的属性表 xlsx。")
with open(attr_path, encoding="utf-8") as f:
    attr = json.load(f)
n_groups = len(attr.get("groups", []))
n_opts = sum(1 for g in attr.get("groups", []) if g.get("opts"))
print(f"[ok] 使用已入库属性表 {attr_path}  (groups={n_groups}, 含opts组={n_opts})，未覆盖")

# ── 2) 规则文本（D 列记号，写入模板 E 列）──
L = []
L.append("@template: Gender")
L.append("")
L.append("# ── 静态字段（gid 不随 Title 长度变化）──")
for n in STATIC:
    L.append(f"[{n}]")
L.append("")
L.append("# ── Title 长度敏感字段（全部用条件固定行，确保输出顺序与模板示例一致）──")
L.append("# 先 ! 忽略，阻止隐式映射重复输出")
L.append("!Name Color")
L.append("!Name")
L.append("!Title")
L.append("")
L.append("# Name Color：仅由 Title 长度决定")
for n in (2, 3, 4, 5, 6):
    L.append(f"? [Title#len:{n}], +[{NAMECOLOR[n]}|Name Color:|[Name Color]];")
L.append("")
L.append("# Name：由 Title 长度 + Name Color 颜色共同决定（共 5×12=60 条）")
for n in (2, 3, 4, 5, 6):
    for i, color in enumerate(NAME_OPTS):
        gid = NAME_BASE[n] + i
        L.append(f"? [Title#len:{n}]&[Name Color:{color}], +[{gid}|{LABEL_FOR[color]}:|[Name]];")
L.append("")
L.append("# Title：仅由 Title 长度决定")
for n in (2, 3, 4, 5, 6):
    L.append(f"? [Title#len:{n}], +[{TITLE[n]}|Title:|[Title]];")
rule_text = "\n".join(L)

# ── 3) 端到端验证 ──
import sys as _sys
_sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import build_callie_rule as bcr
import callie_generator as cg

rule = bcr.compile_rule(SKU, rule_text, "", "", attr)

def run(jtext, tag):
    az, unmatched, _w = cg.generate_callie_v2(jtext, rule, attr, collect_unmatched=True)
    print(f"\n##### {tag} #####")
    print(az)
    if unmatched:
        print("  [unmatched]", unmatched)
    return az

# 样例1：Title=CCHT(4字母)，Name=客人名，Name Color=Dark Purple -> 期望 NameColor=190, Title=191, Name=180(dark purple)
j1 = ("Gender:Woman\nBadge Reel Color:Pink\nSkin Tone:Skin Tone-5\nEyes Color:Eye Color-1\n"
      "Glasses:Eyeglasses-5\nFace Mask:Face Mask-2\nHairstyle:Deepest Blue Black-Hairstyle-2\n"
      "Undershirt:Undershirt-7\nPants:Pants-13\nClothes:Clothes-2\nShoes:Shoes-15\n"
      "Hand-held Item:No\nDrinks:Drinks-13\nName Color:Dark Purple\nName:Marcie\nTitle:CCHT")
az1 = run(j1, "len4 CCHT / Name Color=Dark Purple / Name=Marcie")
assert "190|Name Color:|Dark Purple" in az1, "len4 NameColor 应为 190"
assert "191|Title:|CCHT" in az1, "len4 Title 应为 191"
assert "180|dark purple:|Marcie" in az1, "len4 Name(Dark Purple) 应为 180|dark purple:|Marcie"
print("  [PASS] len4 分支正确")

# 样例2：Title=Rn(2字母)，Name Color=Black，Name=客人名 -> 期望 NameColor=162, Title=163, Name=150(black)
j2 = j1.replace("Name Color:Dark Purple", "Name Color:Black").replace("Name:Marcie", "Name:Alice").replace("Title:CCHT", "Title:Rn")
az2 = run(j2, "len2 Rn / Name Color=Black / Name=Alice")
assert "162|Name Color:|Black" in az2, "len2 NameColor 应为 162"
assert "163|Title:|Rn" in az2, "len2 Title 应为 163"
assert "150|black:|Alice" in az2, "len2 Name(Black) 应为 150|black:|Alice"
print("  [PASS] len2 分支正确")

# 样例3：Title=ICU RN(6字母，含空格)，Name Color=Green，Name=客人名 -> 期望 NameColor=218, Title=219, Name=217(green)
j3 = j1.replace("Name Color:Dark Purple", "Name Color:Green").replace("Name:Marcie", "Name:Bob").replace("Title:CCHT", "Title:ICU RN")
az3 = run(j3, "len6 ICU RN / Name Color=Green / Name=Bob")
assert "218|Name Color:|Green" in az3, "len6 NameColor 应为 218"
assert "219|Title:|ICU RN" in az3, "len6 Title 应为 219"
assert "217|green:|Bob" in az3, "len6 Name(Green) 应为 217|green:|Bob"
print("  [PASS] len6 分支正确")

print("\n[ALL PASS] #len + Name Color 双条件路由端到端通过")

# ── 4) 保存规则草稿文件 ──
draft_path = os.path.join(OUTDIR, "rule_draft.txt")
with open(draft_path, "w", encoding="utf-8") as f:
    f.write(rule_text)
print(f"\n[ok] 规则草稿已保存: {draft_path}")

# ── 5) 注入生产模板（仅当 WRITE_EXCEL=1）──
if os.environ.get("WRITE_EXCEL") == "1":
    XLSX = r"E:\amazon-data\raw\Pet翻译模板.xlsx"
    if os.path.exists(XLSX):
        bak = XLSX + ".bak_caps250123_" + __import__("datetime").datetime.now().strftime("%Y%m%d%H%M%S")
        shutil.copy2(XLSX, bak)
        print(f"[backup] {bak}")
        import openpyxl
        wb = openpyxl.load_workbook(XLSX)
        ws = wb["Sheet1"]
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        sku_c = hdr.index("sku") + 1
        rule_c = hdr.index("callie定制项翻译规则") + 1
        jexample_c = hdr.index("【临时列(需删除)】示例") + 1
        azexample_c = hdr.index("【callie定制项】示例") + 1
        target_row = None
        for row in ws.iter_rows(min_row=2):
            if row[sku_c - 1].value == SKU:
                target_row = row[0].row
                # 读取当前 J 示例，跑引擎生成新的 AZ 示例
                e_text = str(row[jexample_c - 1].value or "")
                sku_cfg = {"version": "2.0", "rule": rule, "attr": attr, "v1": None, "error": None}
                new_f, _unmatched, _w = cg.generate_callie_dispatch(e_text, sku_cfg, collect_unmatched=True)
                row[rule_c - 1].value = rule_text
                row[azexample_c - 1].value = new_f
                print(f"[write] 已写入行 {target_row} 的 callie定制项翻译规则 并刷新 AZ 示例")
                break
        if target_row is None:
            print(f"[warn] 模板中未找到 SKU={SKU}，未写入")
        else:
            # 先保存到临时文件再原子替换，避免文件被占用时 PermissionError
            tmp = XLSX + ".tmp"
            wb.save(tmp)
            try:
                os.replace(tmp, XLSX)
                print("[ok] Pet翻译模板.xlsx 已更新")
            except OSError as e:
                # 文件被其他进程占用时，保留 tmp 文件并提示用户手动替换
                alt = XLSX.replace(".xlsx", f"_{SKU}.xlsx")
                shutil.copy2(tmp, alt)
                os.remove(tmp)
                print(f"[warn] 原文件 {XLSX} 被占用，无法直接覆盖: {e}")
                print(f"[ok] 已生成可替换副本: {alt}")
                print(f"[action] 请关闭占用原文件的程序（如 streamlit / Excel）后，"
                      f"将 {alt} 重命名为 Pet翻译模板.xlsx 覆盖原文件")
