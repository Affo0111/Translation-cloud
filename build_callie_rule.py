# -*- coding: utf-8 -*-
"""
build_callie_rule.py  ——  从「翻译模板（D列备注 + E列J示例 + F列AZ示例）+ 属性表」
自动编译出 callie v2 rule.json。核心引擎（callie_generator.py）零改动，仍吃 rule.json。

流程（用户拍板 2026-08-19）：
  1) 先读属性表（attribute_config.json）→ 学「组ID / 组名 / 选项」
  2) 读翻译模板：E列(J示例) + F列(AZ示例) 反推 field_mapping / key_normalize / 字段模板 / 输出顺序
  3) 读 D列 备注 → 增量覆盖（显式优先）：
        +[ID|:literal]              → fixed_line（固定行，原样输出在最前）
        +[ID|label:|default]        → fixed_field（字段默认值，J 无此字段时补默认）
        !key[:value]                → ignore（抑制该 J 键单独成行，值仍留内存供模板/路由引用）
        [dep1|dep2|..]=[id1|id2|..] → conditional_routing（隐性条件，字段名靠属性表反推）
        ? [cond_field:cond_value] , [dep1|dep2|..]=[id1|id2|..] → conditional_routing 带 when 守卫（条件命中才执行；不匹配则跳过）
  4) 合并 → v2 rule.json
  5) 校验：用 E 跑一遍引擎，输出应 == F，不一致逐行 diff

用法：
  python build_callie_rule.py <SKU> <attr_json> <d.txt> <e.txt> <f.txt> [out.json]
"""

import json
import os
import sys
from collections import OrderedDict

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import callie_generator as cg


# ─────────────────────────────────────────────────────────────────────────────
# 1) D 列备注解析
# ─────────────────────────────────────────────────────────────────────────────

def _split_entries(text):
    """按「分号 / 换行」切分备注条目，去空白；整条以 # 开头的视为注释跳过。"""
    entries = []
    for line in str(text).splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        for seg in line.split(";"):
            seg = seg.strip()
            if seg and not seg.startswith("#"):
                entries.append(seg)
    return entries


def _parse_when(cond_part):
    """把 `? [...]` 的条件部分解析成条件列表（支持 `&` 连接多条件，AND 语义）。

    例：`[Gender:Woman] & [Age:Senior]`
      -> [{"field": "Gender", "value": "Woman"}, {"field": "Age", "value": "Senior"}]
    单条件 `[Gender:Woman]` -> [{"field": "Gender", "value": "Woman"}]。
    始终返回 list（便于引擎统一按 AND 处理，并兼容旧的单 dict 数据）。

    扩展：长度条件 `[Field#len:N]` -> {"field": "Field", "op": "len", "value": "N"}，
    引擎按「字段值的字符长度 == N」判断（用于 Title 字母数分支等自由文本场景）。
    """
    conds = []
    for seg in cond_part.split("&"):
        seg = seg.strip().strip("[]").strip()
        if not seg:
            continue
        if ":" not in seg:
            raise ValueError(f"条件部分应为 [字段:值]: {cond_part!r}")
        f, v = seg.split(":", 1)
        f = f.strip()
        v = v.strip()
        if "#len" in f:
            fname = f.replace("#len", "").strip()
            conds.append({"field": fname, "op": "len", "value": v})
        else:
            conds.append({"field": f, "value": v})
    if not conds:
        raise ValueError(f"条件规则缺少有效条件: {cond_part!r}")
    return conds


def parse_d_notation(d_text):
    """解析 D 列【声明式】备注 → 各结构原始列表。

    支持的新语法（与 A 系统完全隔离，各自解析）：
      +[Field:Default]                      固定字段（J 缺则用默认；ID 由属性表按组名补）
      +[ID|:literal]                        字面固定行（原样输出）
      +[ID|Field:|[JKey]]                   含占位符固定行（[JKey] 从 J 取值）
      ? [Cond:Val] , +[...]                 条件固定行（条件满足才输出该行）
      [Field]                              字段映射（ID 由属性表按组名补）
      [Group:Opt1|Group:Opt2]=[ID1|ID2]     条件映射（隐性路由；Group:Opt 写法）
      ? [Cond:Val] , [Group:Opt]=[ID]       条件映射 + 守卫
      ? [Cond1] & [Cond2] , <动作>          嵌套条件（& = 所有条件同时满足才执行）
      !key(:val)                           忽略字段（不输出自身行，值保留供模板/路由引用）
      # ...                                 注释（整条以 # 开头跳过）
    """
    fixed_fields = []
    fixed_lines = []
    conditional_fixed_lines = []
    field_mapping = []
    conditional_routing = []
    ignore_keys = []
    output_order_fields = []
    template_field = None
    template_default = None

    for raw in _split_entries(d_text):
        if raw.startswith("@template"):
            # @template: Gender —— 声明「模板选择字段」（J 列字段名，其值对应属性表 tmpl）。
            # 引擎运行时读该字段值 → 定位模板 → 字段查 gid 只在「该模板 + public」里找。
            # 支持「@template: Gender = Woman」：等号后为缺省模板（J 无该字段时兜底）。
            body = raw[len("@template"):].lstrip(":").strip()
            if "=" in body:
                _field, _default = body.split("=", 1)
                template_field = _field.strip()
                template_default = _default.strip()
            else:
                template_field = body
            continue
        if raw.startswith("?"):
            # 条件执行：? [条件1] & [条件2] & ... , <动作>（& = 所有条件同时满足）
            body = raw[1:].strip()
            if "," not in body:
                raise ValueError(f"条件规则缺少分隔逗号: {raw!r}")
            cond_part, action = body.split(",", 1)
            when = _parse_when(cond_part)   # list of {field, value}，支持 & 连接
            action = action.strip()
            if action.startswith("+["):
                inner = action[2:-1] if action.endswith("]") else action[2:]
                conditional_fixed_lines.append({"when": when, "raw": inner})
            elif action.startswith("[") and "]=[" in action:
                left, right = action.split("=", 1)
                conditional_routing.append({
                    "dep_raw": [x.strip() for x in left.strip("[]").split("|") if x.strip()],
                    "target_raw": [x.strip() for x in right.strip("[]").split("|") if x.strip()],
                    "when": when,
                })
            else:
                raise ValueError(f"条件规则动作无法识别: {raw!r}")
        elif raw.startswith("+["):
            inner = raw[2:-1] if raw.endswith("]") else raw[2:]
            if "|" not in inner:
                # 固定字段：+[Field:Default]
                if ":" not in inner:
                    raise ValueError(f"固定字段格式应为 +[字段:默认值]: {raw!r}")
                field, default = inner.split(":", 1)
                fixed_fields.append({"field": field.strip(), "default": default.strip()})
                output_order_fields.append(field.strip())
            else:
                # 固定行：字面（+[ID|:literal]）或含占位符（+[ID|Field:|[JKey]]）
                fid, rest = inner.split("|", 1)
                fixed_lines.append(inner)
        elif raw.startswith("!"):
            inner = raw[1:].strip().strip("[]")
            key = inner.split(":", 1)[0].strip()
            if key and key not in ignore_keys:
                ignore_keys.append(key)
        elif raw.startswith("[") and raw.endswith("]") and "]=[" in raw:
            # 条件映射 / 隐性路由：[dep]=[id]（可带 when 守卫在 ? 分支已处理）
            left, right = raw.split("=", 1)
            conditional_routing.append({
                "dep_raw": [x.strip() for x in left.strip("[]").split("|") if x.strip()],
                "target_raw": [x.strip() for x in right.strip("[]").split("|") if x.strip()],
            })
        elif raw.startswith("[") and raw.endswith("]"):
            # 裸字段映射：[Field]
            field = raw[1:-1].strip()
            if field:
                field_mapping.append(field)
                output_order_fields.append(field)
    return {
        "fixed_fields": fixed_fields,
        "fixed_lines": fixed_lines,
        "conditional_fixed_lines": conditional_fixed_lines,
        "field_mapping": field_mapping,
        "conditional_routing": conditional_routing,
        "ignore_keys": ignore_keys,
        "output_order_fields": output_order_fields,
        "template_field": template_field,
        "template_default": template_default,
    }


# ─────────────────────────────────────────────────────────────────────────────
# 2) E 列 / F 列解析
# ─────────────────────────────────────────────────────────────────────────────

def parse_e_column(e_text):
    """解析 E 列 J 示例 → 有序 {key: value}（复用引擎的 parse_j）。"""
    return cg.parse_j(e_text)


def parse_f_column(f_text):
    """解析 F 列 AZ 示例 → 有序 [(kind, id, label, value), ...]。

    kind ∈ {"fixed_line", "mapped"}。格式约定：
      固定行：ID|:literal         （label 为空）
      字段行：ID|label:|value     （含模板，如 ID|Name(Black):|Damary）
    """
    lines = []
    for line in str(f_text).splitlines():
        line = line.strip()
        if not line:
            continue
        id_part, rest = line.split("|", 1)
        if rest.startswith(":"):
            lines.append(("fixed_line", id_part, "", ""))
        else:
            label_part, value = rest.split("|", 1)
            label = label_part.rstrip(":")
            lines.append(("mapped", id_part, label, value))
    return lines


# ─────────────────────────────────────────────────────────────────────────────
# 3) 属性表辅助（反推 conditional_routing 的字段名）
# ─────────────────────────────────────────────────────────────────────────────

def _id_to_names(attr):
    d = {}
    for g in attr.get("groups", []):
        d.setdefault(str(g["gid"]), set()).add(g["name"])
    return d


def _normalize_option(group, value):
    """把 value 归一成该属性组的某个选项（精确 → 忽略大小写 → 前缀）。"""
    opts = group.get("opts", [])
    if value in opts:
        return value
    low = value.lower()
    for o in opts:
        if o.lower() == low:
            return o
    for o in opts:
        ol = o.lower()
        if ol.startswith(low) or low.startswith(ol):
            return o
    return value


def _group_covering_options(attr, values):
    """返回「选项能覆盖 values 里所有值（含模糊归一）」的组；无则 None。"""
    for g in attr.get("groups", []):
        opts = g.get("opts", [])
        if all(_normalize_option(g, v) in opts for v in values):
            return g
    return None


def _find_group_by_name(attr, name):
    """按组名查属性表里的组（兼容 groups_by_name 或直接扫 groups）。"""
    gbn = attr.get("groups_by_name")
    if gbn and name in gbn:
        return gbn[name]
    for g in attr.get("groups", []):
        if g["name"] == name:
            return g
    return None


def _resolve_group_id(attr, name):
    """字段名 → 组 ID。若该字段跨模板(tmpl)有不同 gid（如 Woman=106 / Man=132 / public=174），
    返回 'template' 标记（运行期按模板选择：性别模板优先，public 补充）；
    否则（同模板内多 gid，如基础 gid + 路由目标）返回第一个 gid。"""
    groups = [g for g in attr.get("groups", []) if g["name"] == name]
    if not groups:
        raise ValueError(f"属性表里找不到字段组: {name!r}（请检查属性表或 D 列字段名拼写）")
    tmpl_to_gids = {}
    for g in groups:
        tmpl_to_gids.setdefault(g.get("tmpl", ""), set()).add(str(g["gid"]))
    if len(tmpl_to_gids) > 1:
        return "template"
    return str(groups[0]["gid"])


def _resolve_conditional_routing(cr_list, attr):
    """把 D 列原始 conditional_routing（dep_raw / target_raw）补上
    dependency_field / target_field，产出 v2 格式。

    dep_raw 每项支持两种写法：
      * "Group:Option"  —— 显式组名:选项（推荐，最明确）
      * "Option"        —— 裸选项值（自动找能覆盖它的组）
    """
    id2names = _id_to_names(attr)
    result = []
    for cr in cr_list:
        dep_raw = cr["dep_raw"]
        target_raw = cr["target_raw"]

        # 解析依赖对 (组名, 选项)
        pairs = []
        for d in dep_raw:
            if ":" in d:
                g, o = d.split(":", 1)
                pairs.append((g.strip(), o.strip()))
            else:
                pairs.append((None, d.strip()))

        # ── 漏写组名前缀的容错（常见笔误）──
        # 例：[Beard Color:Black|Beard Color:Dark Brown|Silver]=[143|144|150]
        #     最后一项 Silver 漏了 "Beard Color:" 前缀。
        # 若列表里已有唯一的显式组名，且该裸选项确实是这个组的合法选项，
        # 就自动补上组名（语义无歧义）；不是合法选项才报错，避免静默错配。
        _explicit = [g for g, _ in pairs if g]
        if _explicit and any(g is None for g, _ in pairs):
            _base = _explicit[0]
            if all(g == _base for g in _explicit):
                _base_group = _find_group_by_name(attr, _base)
                if _base_group is not None:
                    _fixed = []
                    for g, o in pairs:
                        if g is None:
                            if _normalize_option(_base_group, o) in _base_group.get("opts", []):
                                g = _base          # 自动补前缀
                            else:
                                raise ValueError(
                                    f"条件映射里 {o!r} 未写组名前缀，且不是 {_base!r} 的合法选项；"
                                    f"请写成 [{_base}:{o}]（合法选项：{_base_group.get('opts')}）")
                        _fixed.append((g, o))
                    pairs = _fixed

        # target_field：所有目标 ID 对应的组名必须唯一
        names = set()
        for tid in target_raw:
            names |= id2names.get(str(tid), set())
        if len(names) != 1:
            raise ValueError(
                f"条件映射目标ID {target_raw} 对应多个组名 {sorted(names)}，无法唯一确定 target_field")
        target_field = names.pop()

        # ── 语义判定 ──
        # rule_field = 左侧显式字段名（如 "Facial Expression"）；None = 裸选项（无冒号）。
        rule_field = pairs[0][0] if pairs and pairs[0][0] else None

        if rule_field is not None and rule_field == target_field:
            # 情况A：字段自身（左侧字段名 == 目标字段）——不是「依赖→目标」路由。
            #   * 选项是字段合法选项 → 固定值映射：[字段:默认值]=[ID]，值=默认值（J 有该字段则覆盖）
            #     （例：[Facial Expression:Open Eyes]=[123] → 输出 123|Facial Expression:|Open Eyes）
            #   * 选项非字段合法选项 → 固定ID：[字段:标签]=[ID]，标签只是上下文，值取 J
            #     （例：[Skin Tone:Young]=[128] → Young 非选项，ID=128，值取 J 的 Skin Tone）
            dep_group = _find_group_by_name(attr, rule_field)
            opts = [o for _, o in pairs]
            is_valid_opt = (dep_group is not None and
                            all(_normalize_option(dep_group, o) in dep_group.get("opts", [])
                                for o in opts))
            if len(target_raw) != 1:
                raise ValueError(
                    f"字段自身映射要求右侧恰好一个ID，实际 {target_raw}（规则 {cr!r}）")
            entry = {
                "dependency_field": None,
                "target_field": target_field,
                "fixed_id": str(target_raw[0]),
            }
            if is_valid_opt and len(pairs) == 1:
                entry["default_value"] = pairs[0][1]
            if cr.get("when"):
                entry["when"] = cr["when"]
            result.append(entry)
            continue

        # 情况B：依赖→目标（rule_field != target_field，或裸选项）。
        #   值映射模式  —— 左值选项是「依赖字段」的有效选项，ID 随 J 该字段的值变
        #                 （例：[Hair Color:Black|...]=[107|...]，Black 真是 Hair Color 选项）
        #   固定ID模式  —— 左值选项「不是」依赖字段的有效选项，ID 由 when 守卫决定。
        if pairs and pairs[0][0] is not None:
            dep_group = _find_group_by_name(attr, pairs[0][0])
            if dep_group is None:
                raise ValueError(f"条件映射依赖组不存在: {pairs[0][0]!r}")
            for g, _ in pairs:
                if g != pairs[0][0]:
                    raise ValueError(f"条件映射依赖组不一致: {pairs}")
            dep_field = dep_group["name"]
            opts = [o for _, o in pairs]
            is_value_map = all(
                _normalize_option(dep_group, o) in dep_group.get("opts", []) for o in opts)
        else:
            opts = [o for _, o in pairs]
            dep_group = _group_covering_options(attr, opts)
            if dep_group is None:
                is_value_map = False
                dep_field = None
            else:
                is_value_map = True
                dep_field = dep_group["name"]

        if is_value_map:
            rules = {}
            for (g, o), tid in zip(pairs, target_raw):
                nv = _normalize_option(dep_group, o)
                rules[nv] = str(tid)
            entry = {
                "dependency_field": dep_field,
                "target_field": target_field,
                "rules": rules,
            }
        else:
            # 固定ID模式：右侧必须恰好一个 ID，值取 J 的 target_field 内容
            if len(target_raw) != 1:
                raise ValueError(
                    f"固定ID模式（左值非字段选项）要求右侧恰好一个ID，实际 {target_raw}（规则 {cr!r}）")
            entry = {
                "dependency_field": None,
                "target_field": target_field,
                "fixed_id": str(target_raw[0]),
            }
        if cr.get("when"):
            entry["when"] = cr["when"]
        result.append(entry)
    return result


# ─────────────────────────────────────────────────────────────────────────────
# 4) E/F 反推 field_mapping + key_normalize
# ─────────────────────────────────────────────────────────────────────────────

def _find_j_key_by_value(e_dict, value):
    """按「值精确匹配」找 J 键；不唯一时返回 None（交给上层兜底）。"""
    hits = [k for k, v in e_dict.items() if v == value]
    if len(hits) == 1:
        return hits[0]
    return None


def infer_mapping(e_dict, f_lines, fixed_field_names):
    """从 E/F 反推 field_mapping 与 key_normalize。

    fixed_field_names：D 列已声明的固定字段名，这些 F 行跳过（不进入 field_mapping）。
    """
    field_mapping = []
    key_normalize = {}
    for kind, fid, label, value in f_lines:
        if kind != "mapped" or label in fixed_field_names:
            continue
        if "(" in label and label.endswith(")"):
            # 模板字段：Name(Black) → field_template "Name([Color])"
            base = label.split("(")[0].strip()
            tmpl_var = label[label.index("(") + 1: label.rindex(")")]
            j_key_base = _find_j_key_by_value(e_dict, value) or base
            j_key_tmpl = _find_j_key_by_value(e_dict, tmpl_var)
            if j_key_tmpl is None:
                raise ValueError(f"模板变量 {tmpl_var!r}（来自 {label}）在 E 列找不到对应字段")
            field_mapping.append({
                "j_key": j_key_base,
                "id": fid,
                "az_field": base,
                "field_template": f"{base}([{j_key_tmpl}])",
            })
        else:
            raw_key = _find_j_key_by_value(e_dict, value) or label
            if raw_key != label:
                key_normalize[raw_key] = label   # J 原始键 → 规范名（引擎先归一再用）
            # j_key 必须是「归一后的键名」（= label），因为引擎按 norm 后的键去 field_mapping 查
            field_mapping.append({"j_key": label, "id": fid, "az_field": label})
    return field_mapping, key_normalize


# ─────────────────────────────────────────────────────────────────────────────
# 5) 编译主入口
# ─────────────────────────────────────────────────────────────────────────────

def compile_rule(sku, d_text, e_text, f_text, attr):
    """把【声明式 D】+ 属性表 编译成 v2 rule.json dict。

    E/F 列仅保留给 validate_rule 做「E 跑引擎 == F」一致性校验；映射以 D 为准。
    """
    d = parse_d_notation(d_text)

    # 路由优先解析，确定哪些字段是 dynamic 目标
    conditional_routing = _resolve_conditional_routing(d["conditional_routing"], attr)
    routed_targets = {r["target_field"] for r in conditional_routing}

    # 固定字段：ID 由属性表按组名补（若该字段同时是 conditional_routing 的目标，
    # 引擎会在「条件命中」时改用路由出的 ID，未命中时回退此基础 ID）。
    fixed_fields = []
    for f in d["fixed_fields"]:
        fid = _resolve_group_id(attr, f["field"])
        fixed_fields.append({"field": f["field"], "id": fid, "default": f["default"]})

    # 固定值映射（[字段:默认值]=[ID]，字段==目标）→ 自动生成固定字段（值=默认，ID 由条件覆盖）。
    # 例：[Facial Expression:Open Eyes]=[123] 会自动等价于
    #     +[Facial Expression:Open Eyes] + ? [..],[Facial Expression]=[123]。
    fixed_value_fields = {}
    for r in conditional_routing:
        if r.get("default_value") is not None:
            fixed_value_fields.setdefault(r["target_field"], r["default_value"])
    existing_fixed_names = {f["field"] for f in fixed_fields}
    for tf, dv in fixed_value_fields.items():
        if tf not in existing_fixed_names:
            base_id = _resolve_group_id(attr, tf)
            fixed_fields.append({"field": tf, "id": base_id, "default": dv})
            existing_fixed_names.add(tf)

    # 字段映射（裸 [Field]）：ID 由属性表按组名补（路由目标标 dynamic）
    field_mapping = []
    for field in d["field_mapping"]:
        fid = "dynamic" if field in routed_targets else _resolve_group_id(attr, field)
        field_mapping.append({"j_key": field, "id": fid, "az_field": field})

    # 输出顺序：D 书写顺序（固定字段 + 字段映射），末尾追加路由目标字段
    output_order = list(d["output_order_fields"]) + [r["target_field"] for r in conditional_routing]
    # 去重保序（固定字段若同时是路由目标，避免在 output_order 出现两次）
    _seen = set()
    output_order = [x for x in output_order if not (x in _seen or _seen.add(x))]

    rule = {
        "config_version": "2.0",
        "sku": sku,
        "fixed_lines": d["fixed_lines"],
        "conditional_fixed_lines": d["conditional_fixed_lines"],
        "fixed_fields": fixed_fields,
        "field_mapping": field_mapping,
        "key_normalize": {},
        "conditional_routing": conditional_routing,
        "value_maps": {},
        "output_order": output_order,
        "ignore_keys": d["ignore_keys"],
        "template_field": d["template_field"],
        "template_default": d["template_default"],
    }
    # 启动期校验（ID 合法性等），提前暴露配置错误
    cg.validate_config_v2(rule, attr)
    return rule


# ─────────────────────────────────────────────────────────────────────────────
# 6) 校验：用 E 跑引擎，对照 F
# ─────────────────────────────────────────────────────────────────────────────

def validate_rule(rule, e_text, f_text, attr):
    """用 E 列跑一遍引擎，对比 F 列，返回 (ok, diffs)。"""
    sku_cfg = {"version": "2.0", "rule": rule, "attr": attr, "v1": None, "error": None}
    try:
        az, unmatched, _warnings = cg.generate_callie_dispatch(e_text, sku_cfg, collect_unmatched=True)
    except cg.CallieError as e:
        return False, [f"引擎报错: {e}"]

    actual = [l.strip() for l in az.splitlines() if l.strip()]
    expected = [l.strip() for l in str(f_text).splitlines() if l.strip()]

    # 归一化：去首尾空格 + 剥掉尾随单个「|」（F 示例写法常有冗余尾 |，与引擎输出无关）
    def _nz(s):
        s = s.strip()
        if s.endswith("|"):
            s = s[:-1]
        return s.strip()

    # 用多重集（计数）比较，而非位置比较：引擎会把 fixed_lines 排在最前、再排字段映射，
    # 行序与 F 示例顺序天然不同；位置式 diff 会误报。集合比较才能正确反映「缺了哪行 / 多了哪行」。
    a_set, e_set = {}, {}
    for l in actual:
        a_set[_nz(l)] = a_set.get(_nz(l), 0) + 1
    for l in expected:
        e_set[_nz(l)] = e_set.get(_nz(l), 0) + 1

    diffs = []
    for l, c in e_set.items():
        gap = c - a_set.get(l, 0)
        for _ in range(gap):
            diffs.append(f"❌ 缺行：{l}")
    for l, c in a_set.items():
        gap = c - e_set.get(l, 0)
        for _ in range(gap):
            diffs.append(f"➕ 多行：{l}")
    return (len(diffs) == 0, diffs)


# ─────────────────────────────────────────────────────────────────────────────
# 6.5) 从单条 E/F 示例 反推 D 列规则【草稿】文本（系统生成 0.5，运营改到 1.0）
# ─────────────────────────────────────────────────────────────────────────────

def _norm_key(s):
    """归一化键名：小写、去首尾空格、全角冒号转半角、去尾部冒号。"""
    return str(s).strip().lower().replace("：", ":").rstrip(":").strip()


def _attr_name_gids(attr, name):
    """返回属性表里某个字段名对应的所有 gid（多 gid 组如 Beard Style 会返回多个）。"""
    n = _norm_key(name)
    gids = set()
    for g in (attr or {}).get("groups", []):
        if _norm_key(g.get("name", "")) == n:
            gids.add(str(g["gid"]))
    return gids


def _parse_f_lines(f_text):
    """解析 F 列 AZ 示例 → [(kind, id, label, value), ...]。

    kind ∈ {"fixed_line", "mapped"}。format：
      固定行：ID|:literal         （label 为空，literal 可能含尾随 |）
      字段行：ID|label:|value     （label 为字段名，value 为值）
    """
    out = []
    for line in str(f_text).splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "|" not in line:
            continue
        id_part, rest = line.split("|", 1)
        id_part = id_part.strip()
        if rest.startswith(":"):
            literal = rest[1:].rstrip("|").strip()
            out.append(("fixed_line", id_part, "", literal))
        else:
            label_part, value = rest.split("|", 1)
            label = label_part.rstrip(":").strip()
            out.append(("mapped", id_part, label, value.strip()))
    return out


def _detect_template_field(attr):
    """从属性表推断「模板选择字段」：其选项集合 == 非 public 的 tmpl 维度值集合。

    例：PG5934 的属性表 tmpl 维度 = {Female, Male, Kid}，而 Character 的 opts={Female,Male,Kid}
        → 返回 'Character'（即 J 里用 Character 的值来选属性表模板）。
    PL251368 的 tmpl 维度 = {Woman, Man}，Gender 的 opts={Woman,Man} → 返回 'Gender'。
    找不到（属性表无模板拆分）则返回 None。
    """
    if not attr:
        return None
    groups = attr.get("groups", [])
    tmpls = set(str(g.get("tmpl", "")) for g in groups) - {"", "public"}
    if not tmpls:
        return None
    tnorm = {t.strip().lower() for t in tmpls}
    name_opts = {}
    for g in groups:
        name = g.get("name", "")
        if not name:
            continue
        for o in g.get("opts", []):
            name_opts.setdefault(name, set()).add(str(o).strip().lower())
    for name, opts in name_opts.items():
        if opts and opts == tnorm:
            return name
    return None


def generate_rule_draft(e_text, f_text, attr=None):
    """从单条 J 示例(E) + AZ 示例(F) 反推 D 列规则【草稿】文本。

    定位：系统生成 0.5，运营审核改到 1.0。仅覆盖机械部分：
      - 平铺字段映射 [Field]（J 与 AZ 同名字段且值一致；单 gid 字段，引擎按属性表补 gid）
      - 模板拆分字段（如 Skin Tone 在性别/年龄下各有多 gid）：锁死样本里的 gid
        （引擎在 fixed_fields 之外无法把 [Field] 解析成具体数字 gid，会输出非法 template|...，故必须写死）
      - 占位符固定行 +[gid|Field:|[src]]（AZ 值取自「另一个 J 字段」，gid 引擎硬要求）
      - 字面固定行 +[gid|Field:|value] / +[gid|:literal]
      - @template 探测（从属性表 tmpl 维度推断模板字段，而非仅看 J 是否含 Gender）
    每条语句以 ';' 结尾（@template / # 注释除外），对齐手写规范。

    关键约束：绝不使用 [Dep]=[gid] 值映射——目标字段不在 J 列时会静默出空行。
    跨字段取值一律用「固定行 + [占位符]」表达（值来自其它 J 字段）。
    """
    e_dict = parse_e_column(e_text) or {}
    if not e_dict:
        return "# ⚠️ E 列（J 示例）为空，无法生成草稿"

    # 归一化 J 键；同时建「值 -> [J键]」索引（用于占位符判定）
    e_norm = {_norm_key(k): v for k, v in e_dict.items()}
    val_to_keys = {}
    for k, v in e_dict.items():
        val_to_keys.setdefault(v, []).append(k)

    self_mapped = set()  # 已用 [Field] 自映射表示的 J 字段（不需 ! 抑制）

    # 手写习惯分组桶：@template → 固定行 → 字段映射 → 条件/路由(预留) → 忽略
    template_lines = []
    fixed_lines = []       # 所有 +[...] 结构行（字面固定 / 固定行 / 占位符取自 J）
    mapping_lines = []     # 显式 [Field]（仅 attr 缺失降级时产生）
    conditional_lines = []  # ? [...] 条件/路由（草稿当前不生成，预留）
    ignore_lines = []      # ![field] 忽略未使用字段（放最后做清理）

    # @template 探测：从属性表推断模板字段（组的 tmpl 维度对应的 J 字段）
    tmpl_field = _detect_template_field(attr)
    if tmpl_field:
        template_lines.append(f"@template: {tmpl_field}")

    def _stmt(s):
        if s.startswith("@template") or s.lstrip().startswith("#"):
            return s
        return s if s.endswith(";") else s + ";"

    def _gid_for(field, fallback_fid):
        # 占位符/字面固定行的 gid 由属性表按字段名补；补不到（如字段名带空格差异）则用 F 的 gid
        if not attr:
            return fallback_fid
        try:
            g = _resolve_group_id(attr, field)
        except ValueError:
            return fallback_fid
        return fallback_fid if (g is None or g == "template") else g

    # 逐 AZ 行分类
    for kind, fid, label, value in _parse_f_lines(f_text):
        if kind == "fixed_line":
            fixed_lines.append(_stmt(f"+[{fid}|:{value}]"))
            continue
        field = (label or "").strip()
        if not field:
            fixed_lines.append(_stmt(f"+[{fid}|:{value}]"))
            continue
        nf = _norm_key(field)
        # 1) 自映射：J 有同名字段且值一致
        if nf in e_norm and e_norm[nf] == value:
            if attr is not None and len(_attr_name_gids(attr, field)) > 1:
                # 多 gid 字段（含「模板拆分字段」如 Skin Tone，或同模板多 gid 如 Gender）：
                # 引擎无法用 [Field] 稳定解析出与 F 一致的 gid（模板拆分会输出非法 template|...，
                # 同模板多 gid 会取首个而非 F 所用），故锁死样本里的 gid；
                # 同时不加入 self_mapped → 下方 ! 循环补 !{field} 抑制隐式映射重复出。
                fixed_lines.append(_stmt(f"+[{fid}|{field}:|[{field}]]"))
            else:
                # 单 gid 自映射：引擎「隐式字段映射」会自动查属性表 name→gid 补 ID 输出，
                # 无需显式写 [Field]（冗余且增加审核负担）。仅标记 self_mapped 防止下方 ! 循环
                # 误抑制该字段（误抑制会让隐式映射也不输出）。
                # 降级：无属性表(attr is None)时无法依赖隐式映射，仍显式写出以便预览。
                if attr is None:
                    mapping_lines.append(_stmt(f"[{field}]"))
                self_mapped.add(nf)
            continue
        # 2) 值取自另一个 J 字段 → 占位符固定行（gid 引擎硬要求，按属性表补）
        if value in val_to_keys:
            src = val_to_keys[value][0]
            fixed_lines.append(_stmt(f"+[{_gid_for(field, fid)}|{field}:|[{src}]]"))
            continue
        # 3) 纯字面固定行
        fixed_lines.append(_stmt(f"+[{_gid_for(field, fid)}|{field}:|{value}]"))

    # 4) 对每个「非自映射」的 J 字段加 ! 抑制隐式映射
    #    （路由/条件键、被固定行消费的字段不应单独成行）
    for k in e_dict:
        if _norm_key(k) not in self_mapped:
            ignore_lines.append(_stmt(f"!{k}"))

    # 手写习惯顺序拼接：@template → 固定行 → 字段映射 → 条件/路由 → 忽略
    out = []
    out.extend(template_lines)
    if fixed_lines:
        out.append("# 固定行")
        out.extend(fixed_lines)
    if mapping_lines:
        out.append("# 字段映射")
        out.extend(mapping_lines)
    if conditional_lines:
        out.append("# 条件/路由")
        out.extend(conditional_lines)
    if ignore_lines:
        out.append("# 忽略未使用字段")
        out.extend(ignore_lines)
    return "\n".join(out)


# ─────────────────────────────────────────────────────────────────────────────
# 7) 从翻译模板 / 属性表 批量编译 + 订单处理（供 UI 调用）
# ─────────────────────────────────────────────────────────────────────────────

def _detect_columns(template_path):
    """按表头名定位关键列，返回 {role: 0-based 列号}。

    兼容 Pet 翻译模板（sku 在第 B 列）与 Supply 翻译模板（前面多一个「序号」列，
    sku 在第 C 列）两种新增「标准化规则」列后的布局；并对极旧布局做兜底。

    role 含义：
      sku      —— 产品 SKU 列
      std      —— 标准化规则列（新增，可能为空/缺失）
      mapping  —— callie 定制项翻译规则列
      j        —— 【临时列(需删除)】示例列
      az       —— 【callie定制项】示例列
      product  —— calie商品ID和商品版本列
    """
    import openpyxl
    wb = openpyxl.load_workbook(template_path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}
    header = [str(h or "").strip() for h in rows[0]]
    n = len(header)

    def first(pred):
        for i, h in enumerate(header):
            if pred(h):
                return i
        return None

    def has(*subs):
        return lambda h: all(s.lower() in h.lower() for s in subs)

    cols = {}
    # sku：精确匹配 "sku"（Pet=B1、Supply=C1 均为小写的 sku）
    sku = first(lambda h: h.strip().lower() == "sku")
    if sku is None:
        sku = 1 if n > 1 else 0
    cols["sku"] = sku

    # 标准化规则：含 "标准化规则"（缺失则为 None）
    cols["std"] = first(has("标准化规则"))

    # 映射规则：优先 "callie定制项翻译规则"，兜底 "翻译规则"，再兜底第 D 列(旧布局)
    mapping = first(has("callie定制项翻译规则"))
    if mapping is None:
        mapping = first(has("翻译规则"))
    if mapping is None and n > 3:
        mapping = 3
    cols["mapping"] = mapping

    # J 示例：含 "临时列" 或 "需删除"
    j = first(lambda h: ("临时列" in h) or ("需删除" in h))
    if j is None and n > 4:
        j = 4
    cols["j"] = j

    # AZ 示例：同时含 "callie定制项" 与 "示例"（避免与映射规则列冲突）
    az = first(has("callie定制项", "示例"))
    if az is None and n > 5:
        az = 5
    cols["az"] = az

    # 商品信息：含 "商品" 且 ("版本" 或 "calie商品id")
    product = first(lambda h: ("商品" in h) and (("版本" in h) or ("calie商品id" in h.lower())))
    if product is None and n > 6:
        product = 6
    cols["product"] = product

    return cols


def read_template(template_path):
    """读取翻译模板 → {sku: {"D":映射规则, "E":J示例, "F":AZ示例, "STD":标准化规则}}。

    按表头名定位列（兼容 Pet / Supply 两种布局及极旧布局兜底）。
    只收录 映射/J/AZ/标准化 至少一列非空的 SKU（其余视为无需 callie 定制）。
    """
    import openpyxl
    wb = openpyxl.load_workbook(template_path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    cols = _detect_columns(template_path)
    sku_c, map_c, j_c, az_c, std_c = (cols["sku"], cols["mapping"],
                                      cols["j"], cols["az"], cols["std"])

    def _cell(r, c):
        if c is None or c >= len(r):
            return ""
        return str(r[c] or "").strip()

    result = {}
    for r in rows[1:]:
        sku = _cell(r, sku_c)
        if not sku:
            continue
        d = _cell(r, map_c)
        e = _cell(r, j_c)
        f = _cell(r, az_c)
        std = _cell(r, std_c)
        if not (d or e or f or std):
            continue
        result[sku] = {"D": d, "E": e, "F": f, "STD": std}
    return result


def compile_from_template(template_path, attr_sources):
    """从翻译模板 + 属性表，批量编译所有 SKU 的规则。

    attr_sources: {sku: attr}，attr 可以是 attribute.xlsx 路径，或已编译的 attr dict。
    返回 {sku: sku_cfg}，sku_cfg 结构同 load_sku_configs 的 value：{version, rule, attr, v1, error}。
    """
    template = read_template(template_path)
    cfgs = {}
    for sku, cols in template.items():
        attr = attr_sources.get(sku)
        std_text = cols.get("STD", "")
        if attr is None:
            cfgs[sku] = {"version": "2.0", "rule": None, "attr": None, "v1": None,
                         "error": f"缺少属性表（未上传 {sku}.xlsx）", "std_text": std_text}
            continue
        if isinstance(attr, str):
            if attr.endswith(".json"):
                with open(attr, encoding="utf-8") as f:
                    attr = json.load(f)
            else:
                attr = cg.compile_attribute(attr)
        try:
            rule = compile_rule(sku, cols["D"], cols["E"], cols["F"], attr)
            cfgs[sku] = {"version": "2.0", "rule": rule, "attr": attr, "v1": None,
                         "error": None, "std_text": std_text}
        except Exception as e:
            cfgs[sku] = {"version": "2.0", "rule": None, "attr": attr, "v1": None,
                         "error": str(e), "std_text": std_text}
    return cfgs


def parse_callie_product(g_text):
    """把翻译模板 G 列文本解析成 (callie_id, callie_version)。

    约定（每行一条，冒号分隔，兼容全角/半角冒号）：
        callie商品id：30694
        callie商品版本：1.2.48.260812-C40
    找不到对应键时返回 (None, None)。
    """
    cid = None
    cver = None
    for line in str(g_text).splitlines():
        line = line.strip()
        if not line:
            continue
        if "：" in line:
            key, _, val = line.partition("：")
        elif ":" in line:
            key, _, val = line.partition(":")
        else:
            continue
        key = key.strip()
        val = val.strip().rstrip(";")  # 去掉规则分号残留
        kl = key.lower()
        if "商品id" in kl:
            cid = val
        elif "商品版本" in kl:
            cver = val
    return cid, cver


def _norm_sku(sku):
    """SKU 归一化：去首尾空格、转小写，用于模板商品信息与订单 SKU 的稳健匹配。

    订单里的 SKU 常带多余空格或大小写差异，导致模板「calie商品ID和商品版本」
    列里配好的商品信息匹配不上，AW/AX/AY 三列因此随机留空。归一化后稳定命中。
    """
    return str(sku or "").strip().lower()


def build_callie_product_map(template_path):
    """读取翻译模板，返回 {sku: {"id":..., "version":...}}（仅含 G 列非空的 SKU）。

    与 read_template 解耦：这里扫描整张表的所有 SKU（即使该 SKU 没有 callie 定制项规则），
    只要 G 列（calie商品ID和商品版本）非空就收录，供 B 系统导出时填充
    AW(参考callie站点)/AX(calie商品ID)/AY(callie商品版本) 三列。
    """
    import openpyxl
    wb = openpyxl.load_workbook(template_path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    cols = _detect_columns(template_path)
    sku_c, prod_c = cols["sku"], cols["product"]

    def _cell(r, c):
        if c is None or c >= len(r):
            return ""
        return str(r[c] or "").strip()

    result = {}
    for r in rows[1:]:
        sku = _cell(r, sku_c)
        if not sku:
            continue
        g = _cell(r, prod_c)
        if not g:
            continue
        cid, cver = parse_callie_product(g)
        if cid or cver:
            result[_norm_sku(sku)] = {"id": cid, "version": cver}
    return result


def process_orders(df, sku_cfgs, callie_product_map=None,
                   sku_col="SKU", j_col="临时列(需删除)", az_col="callie定制项",
                   site_col="参考callie站点", id_col="calie商品ID", ver_col="callie商品版本"):
    """对订单 DataFrame 逐行：生成 callie 定制项（AZ）并填充 callie 商品信息三列
    （AW=参考callie站点 / AX=calie商品ID / AY=callie商品版本，以原始订单 pet8.17.xls 为准）。

    callie 定制项来自 D 列规则编译；三列商品信息来自翻译模板 G 列
    （按 SKU 匹配，经 callie_product_map 传入；只有该 SKU 在模板 G 列配了信息，
    AW/AX/AY 三列才会被填充，否则留空）。
    返回 (result_df, errors)。errors 每项：{行号, SKU, 状态, 说明}。
    """
    import pandas as pd
    result = df.copy()
    for c in (az_col, site_col, id_col, ver_col):
        if c not in result.columns:
            result[c] = ""
    callie_product_map = callie_product_map or {}
    # 归一化 SKU 键，避免订单 SKU 与模板/商品信息 SKU 的大小写/空格差异导致匹配不上
    _sku_cfgs = {_norm_sku(k): v for k, v in (sku_cfgs or {}).items()}
    _cp_map = {_norm_sku(k): v for k, v in callie_product_map.items()}
    errors = []
    _warned_no_product = set()
    for idx, row in result.iterrows():
        sku = str(row.get(sku_col, "") or "").strip()
        j = str(row.get(j_col, "") or "").strip()
        if not sku and not j:
            continue
        line_no = (idx + 2) if isinstance(idx, int) else ""
        sku_n = _norm_sku(sku)
        cfg = _sku_cfgs.get(sku_n)
        cp = _cp_map.get(sku_n)

        # ── callie 商品信息三列（AW=参考callie站点 / AX=calie商品ID / AY=callie商品版本）──
        # 三列均来自翻译模板 G 列：只有该 SKU 在模板里配了 callie 商品信息，才填充 AW=callie / AX / AY；
        # 未配 G 的 SKU 全部留空（与 A 系统「无模板 = 不翻译」的设计一致）。
        if cp:
            result.at[idx, site_col] = "callie"
            result.at[idx, id_col] = cp.get("id") or ""
            result.at[idx, ver_col] = cp.get("version") or ""
        elif cfg is not None and not sku_n in _warned_no_product:
            # AZ 能生成但没配商品信息 → 明确提示，避免 AW/AX/AY「随机留空」却看不到原因
            _warned_no_product.add(sku_n)
            errors.append({"行号": "", "SKU": sku, "状态": "无商品信息",
                           "说明": "翻译模板「calie商品ID和商品版本」列未配该 SKU，AW/AX/AY 留空"})

        # ── callie 定制项（AZ / az_col）：来自 D 列规则编译 ──
        if cfg is None:
            if cp is None:
                errors.append({"行号": line_no, "SKU": sku, "状态": "未配置",
                               "说明": "翻译模板无此 SKU 的 callie 配置"})
            continue
        # ⚠️ 顺序要紧：先报「配置错误」，再判 rule 是否存在。
        # 否则规则编译失败（rule=None + error 非空）会被「无 rule 就静默跳过」吞掉，
        # 表现为 AZ 整列空、错误报告里却看不到任何原因（2026-08-21 实际踩坑）。
        if cfg.get("error"):
            errors.append({"行号": line_no, "SKU": sku, "状态": "配置错误", "说明": cfg["error"]})
            continue
        if not cfg.get("rule"):
            errors.append({"行号": line_no, "SKU": sku, "状态": "配置错误",
                           "说明": "该 SKU 规则未能编译出结果（D 列可能为空或语法无法识别）"})
            continue
        # ── 阶段 0：J 列标准化前置（在规则执行前清洗脏数据）──
        # 该 SKU 的标准化规则来自翻译模板「标准化规则」列（read_template 已读入 cfg["std_text"]）。
        std_text = cfg.get("std_text", "")
        if std_text:
            std_rules = cg.parse_std_rules(std_text)
            if std_rules:
                j = cg.normalize_j(j, std_rules)
        try:
            az, unmatched, warnings = cg.generate_callie_dispatch(j, cfg, collect_unmatched=True)
            result.at[idx, az_col] = az
            if unmatched:
                errors.append({"行号": line_no, "SKU": sku, "状态": "部分未匹配",
                               "说明": "、".join(u["field"] for u in unmatched)})
            for w in warnings:
                errors.append({"行号": line_no, "SKU": sku, "状态": "占位符缺失警告",
                               "说明": f"字段 {w['field']} 在J列缺失，已输出空值行"})
        except cg.CallieError as e:
            errors.append({"行号": line_no, "SKU": sku, "状态": "生成失败", "说明": str(e)})
        except Exception as e:
            errors.append({"行号": line_no, "SKU": sku, "状态": "异常", "说明": str(e)})
    return result, errors


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def _read(p):
    with open(p, encoding="utf-8") as f:
        return f.read()


def main(sku=None, attr_json=None, d_path=None, e_path=None, f_path=None, out_path=None):
    if sku is None:  # CLI 模式
        if len(sys.argv) < 6:
            print("用法: python build_callie_rule.py <SKU> <attr.json> <d.txt> <e.txt> <f.txt> [out.json]")
            sys.exit(1)
        sku, attr_json = sys.argv[1], sys.argv[2]
        d_path, e_path, f_path = sys.argv[3], sys.argv[4], sys.argv[5]
        out_path = sys.argv[6] if len(sys.argv) > 6 else None

    with open(attr_json, encoding="utf-8") as f:
        attr = json.load(f)
    rule = compile_rule(sku, _read(d_path), _read(e_path), _read(f_path), attr)
    ok, diffs = validate_rule(rule, _read(e_path), _read(f_path), attr)

    if out_path is None:
        base = os.path.join(os.path.dirname(os.path.abspath(__file__)), "callie_sku_configs", sku)
        out_path = os.path.join(base, "rule.json")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(rule, f, ensure_ascii=False, indent=2)

    print(f"[编译] {sku} -> {out_path}")
    print(f"[校验] {'✅ 通过：E 列跑引擎输出 == F 列' if ok else '❌ 不一致：'}")
    for d_ in diffs:
        print("   ", d_)
    return ok


if __name__ == "__main__":
    main()
