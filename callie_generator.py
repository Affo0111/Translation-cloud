# -*- coding: utf-8 -*-
"""
callie_generator.py  ——  callie 定制项【通用引擎】

设计目标（用户硬性要求）：
  核心代码【不硬编码】任何 SKU 专属信息（角色名、字段名、字段顺序、
  属性 ID、映射关系、固定前缀、Style 别名等）。每一个 SKU 的全部语义
  都由两份【每 SKU 独立的配置文件】驱动：
    1) attribute.xlsx  —— 该 SKU 的属性表（组/选项/ID），给出"有哪些"。
    2) rule.json       —— 该 SKU 的规则配置（用户上传），给出"怎么解释"。
  换 SKU 时只需替换这两份文件，核心代码一行不动。

模块职责：
  * compile_attribute(xlsx)         —— 把 attribute.xlsx 解析成与 SKU 无关的中间结构
  * load_rule_config(json)          —— 读取用户上传的规则配置
  * build_runtime_config(attr, rule)—— 把两者合并成运行时配置（纯数据，无字面量）
  * generate_callie(j_text, cfg)    —— 纯函数：J 列文本 + 运行时配置 -> AZ 列文本

所有"角色/字段/ID/路由"都来自 cfg，代码里没有任何 CAPS241501 的字面量。
"""

import json
import os
import re
from collections import OrderedDict


# ─────────────────────────────────────────────────────────────────────────────
# 异常
# ─────────────────────────────────────────────────────────────────────────────

class CallieError(Exception):
    """callie 生成过程中的可恢复异常（未知Key / 未知角色 / 未知发色等）。"""
    pass


# ─────────────────────────────────────────────────────────────────────────────
# 1) 解析 J 列
# ─────────────────────────────────────────────────────────────────────────────

def parse_j(text):
    """
    将 J 列多行文本解析为有序键值对 dict。
    规则：按行分割；每行按第一个 ':' 拆成 key/value；重复 key 以【第一次】为准
    （背景/基础值优先，符合翻译模板 F 列示例，如 PL251368 的 Color 取首个 Dark Blue/ Green）。
    空行、无 ':' 的行跳过。
    """
    data = OrderedDict()
    if not text:
        return data
    for line in str(text).splitlines():
        line = line.rstrip("\r")
        if not line.strip():
            continue
        if ":" not in line:
            continue
        key, val = line.split(":", 1)
        key = key.strip()
        val = val.strip()
        if key:
            if key not in data:
                data[key] = val  # 重复 key -> 保留第一次（背景/基础值）
    return data


# ─────────────────────────────────────────────────────────────────────────────
# 2) 解析 attribute.xlsx（与 SKU 无关的中间结构）
# ─────────────────────────────────────────────────────────────────────────────

def compile_attribute(xlsx_path):
    """
    读取 attribute.xlsx 的 'data' 表，解析成与具体 SKU 无关的中间结构：

      {
        "groups": [ {name, gid, opts:[...], tmpl}, ... ],   # 全部属性组
        "groups_by_name": { name: {gid, opts, tmpl}, ... }, # 名字 -> 组
      }

    列约定（0-based）：1=模板代码, 2=模板名称, 3=属性组id, 4=属性组名称, 5=属性id, 6=名称(选项)
    这里【不假设】任何字段叫什么名字，只是把"组/选项/ID"原样读出来。
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["data"] if "data" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))

    groups = []
    by_name = OrderedDict()

    for r in rows[1:]:
        if len(r) < 7 or r[3] is None:
            continue
        gid = r[3]
        gname = (r[4] or "").strip()
        opt = r[6]
        tmpl = (r[2] or "").strip()
        rec = None
        for g in groups:
            if g["gid"] == gid and g["name"] == gname:
                rec = g
                break
        if rec is None:
            rec = {"name": gname, "gid": gid, "opts": [], "tmpl": tmpl}
            groups.append(rec)
            by_name[gname] = rec
        if opt is not None and str(opt).strip():
            rec["opts"].append(str(opt).strip())

    return {"groups": groups, "groups_by_name": by_name}


def save_attribute_config(xlsx_path, json_path):
    """把 attribute.xlsx 编译成 attribute_config.json（运行时无需 openpyxl）。"""
    attr = compile_attribute(xlsx_path)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(attr, f, ensure_ascii=False, indent=2)
    return attr


# ─────────────────────────────────────────────────────────────────────────────
# 3) 读取规则配置（用户上传，每 SKU 一份）
# ─────────────────────────────────────────────────────────────────────────────

def load_rule_config(json_path):
    with open(json_path, encoding="utf-8") as f:
        return json.load(f)


# ─────────────────────────────────────────────────────────────────────────────
# 4) 合并成运行时配置（核心：这里【没有任何 SKU 字面量】）
# ─────────────────────────────────────────────────────────────────────────────

def _find_groups_by_name(attr, name):
    """返回所有名为 name 的组（路由组会有多个同名组）。"""
    return [g for g in attr["groups"] if g["name"] == name]


def _zip_pad(keys, vals):
    """把「选项列表」与「组ID列表」按位置对齐成 dict。

    语义：第 i 个选项 -> 第 i 个组ID。当两者长度不一致时（不同 SKU 的属性表
    可能 N 个发色对应 1 个发型组，或反过来），用较短列表的【最后一个元素】补齐
    较长列表的剩余项，绝不让任何选项落空（否则会触发「未知发色」报错）。
    例：colors=[White,Black], sgids=[703] -> {White:703, Black:703}
    """
    if not keys or not vals:
        return {}
    last = vals[-1]
    return {k: vals[min(i, len(vals) - 1)] for i, k in enumerate(keys)}


def build_runtime_config(attr, rule):
    """
    把 attribute 中间结构 + 用户规则配置，合并成 generate_callie 用的运行时配置。

    rule.json 结构（全部由用户按 SKU 填写，引擎不写死任何内容）：
      {
        "sku": "CAPS241501",
        "fixed_prefix": ["1|:10001|", "172|:1720001|"],          # 每 SKU 不同
        "filter_keys": ["font", "font color"],                  # 丢弃的键（小写）
        "character_field": "Character",                         # 角色选择组名
        "style_field": {                                        # 全局样式组
            "name": "Style",
            "aliases": { "缩写写法": "标准写法", ... }           # 可选；缺省则不映射
        },
        "shared_fields": ["Birth Flower", "Name"],             # 全局（非角色）字段
        "hair_routing":  { "color_field": "Hair Color",  "style_field": "Hair Style" },
        "beard_routing": { "color_field": "Beard Color", "style_field": "Beard Style" },
        "no_hair":  {"field": "No Hair",  "option": "Bald"},    # 可选
        "no_beard": {"field": "No Beard", "option": "No"}       # 可选
      }
    """
    char_name = rule["character_field"]
    char_g = attr["groups_by_name"].get(char_name)
    if char_g is None:
        raise CallieError(f"规则配置指定的角色组不存在于属性表: {char_name!r}")
    roles = list(char_g["opts"])
    char_gid = char_g["gid"]

    style_name = rule["style_field"]["name"]
    style_g = attr["groups_by_name"].get(style_name)
    if style_g is None:
        raise CallieError(f"规则配置指定的样式组不存在于属性表: {style_name!r}")
    style_gid = style_g["gid"]
    _sf = rule["style_field"]
    style_aliases = _sf.get("aliases") or {}          # 方案3：精确对照
    style_alias_rules = _sf.get("alias_rules") or []    # 方案1：关键词组合

    # 共享（全局）字段
    shared = {}
    for sf in rule.get("shared_fields", []):
        sg = attr["groups_by_name"].get(sf)
        if sg is None:
            raise CallieError(f"规则配置指定的共享字段不存在于属性表: {sf!r}")
        shared[sf] = sg["gid"]

    # 收集"路由样式组"的 gid，避免它们进入普通名字->gid 映射（同名多组）
    routed_style_keys = set()
    routed_beard_keys = set()
    routing = {}       # role -> {color_option: style_gid}
    beard_routing = {}  # role -> {color_option: beard_gid}
    no_hair = {}        # role -> (gid, option)
    no_beard = {}       # role -> (gid, option)

    hr = rule.get("hair_routing")
    if hr:
        cf, sf = hr["color_field"], hr["style_field"]
        for role in roles:
            cg = attr["groups_by_name"].get(f"{role}'s {cf}")
            sgids = sorted(g["gid"] for g in _find_groups_by_name(attr, f"{role}'s {sf}"))
            if cg is None or not sgids:
                continue
            colors = list(cg["opts"])
            nh = rule.get("no_hair")
            if nh:
                nh_group = attr["groups_by_name"].get(nh["field"])
                if nh_group is not None:
                    nh_opt = nh.get("option")
                    no_hair[role] = (nh_group["gid"], nh_opt)
                    if nh_opt and nh_opt in colors:
                        colors = [c for c in colors if c != nh_opt]
            routing[role] = _zip_pad(colors, sgids)
            routed_style_keys.add(f"{role}'s {sf}")

    br = rule.get("beard_routing")
    if br:
        cf, sf = br["color_field"], br["style_field"]
        for role in roles:
            cg = attr["groups_by_name"].get(f"{role}'s {cf}")
            sgids = sorted(g["gid"] for g in _find_groups_by_name(attr, f"{role}'s {sf}"))
            if cg is None or not sgids:
                continue
            colors = list(cg["opts"])
            nb = rule.get("no_beard")
            if nb:
                nb_group = attr["groups_by_name"].get(nb["field"])
                if nb_group is not None:
                    nb_opt = nb.get("option")
                    no_beard[role] = (nb_group["gid"], nb_opt)
                    if nb_opt and nb_opt in colors:
                        colors = [c for c in colors if c != nb_opt]
            beard_routing[role] = _zip_pad(colors, sgids)
            routed_beard_keys.add(f"{role}'s {sf}")

    # 普通字段：名字 -> gid（路由样式/胡子组已排除）
    groups = {}
    for g in attr["groups"]:
        if g["name"] in routed_style_keys or g["name"] in routed_beard_keys:
            continue
        groups.setdefault(g["name"], g["gid"])

    # 合法键集合（用于未知 Key 校验）
    valid_keys = set(groups.keys()) | routed_style_keys | routed_beard_keys
    valid_keys.add(char_name)
    valid_keys.add(style_name)

    hair_style_field = hr["style_field"] if hr else None
    beard_style_field = br["style_field"] if br else None

    cfg = {
        "character_field": char_name,
        "char_gid": char_gid,
        "roles": roles,
        "style_field": style_name,
        "style_gid": style_gid,
        "style_aliases": style_aliases,
        "shared": shared,
        "groups": groups,
        "routing": routing,
        "beard_routing": beard_routing,
        "no_hair": no_hair,
        "no_beard": no_beard,
        "hair_style_field": (hr["style_field"] if hr else None),
        "beard_style_field": (br["style_field"] if br else None),
        "valid_keys": valid_keys,
        "fixed_prefix": list(rule.get("fixed_prefix", [])),
        "filter_keys": set(k.lower() for k in rule.get("filter_keys", [])),
        # 供 generate 直接取用的颜色字段名 / no_hair / no_beard 字段名
        "_color_field_hair": (hr["color_field"] if hr else None),
        "_color_field_beard": (br["color_field"] if br else None),
        "_no_hair_field": ((rule.get("no_hair") or {}).get("field")),
        "_no_beard_field": ((rule.get("no_beard") or {}).get("field")),
        "style_alias_exact": style_aliases,
        "style_alias_rules": style_alias_rules,
        "output_order": (
            rule["output_order"]
            if rule.get("output_order")
            else _default_output_order(
                roles, style_name, char_name, hair_style_field, beard_style_field,
                attr["groups"], routed_style_keys, routed_beard_keys,
            )
        ),
    }
    return cfg


def load_sku_configs(base_dir):
    """
    扫描 base_dir 下每个子目录（以 SKU 命名），读取其中的 rule.json +
    attribute.xlsx / attribute_config.json，构建 {sku: runtime_cfg}。
    换 SKU 只需把对应子目录换掉，核心代码不动。
    """
    cfgs = {}
    if not os.path.isdir(base_dir):
        return cfgs
    for key in sorted(os.listdir(base_dir)):
        d = os.path.join(base_dir, key)
        if not os.path.isdir(d):
            continue
        rj = os.path.join(d, "rule.json")
        if not os.path.isfile(rj):
            continue
        rule = load_rule_config(rj)
        attr = None
        aj = os.path.join(d, "attribute_config.json")
        ax = os.path.join(d, "attribute.xlsx")
        try:
            if os.path.isfile(aj):
                with open(aj, encoding="utf-8") as f:
                    attr = json.load(f)
            elif os.path.isfile(ax):
                attr = compile_attribute(ax)
        except Exception as e:
            raise CallieError(f"SKU {key!r} 属性表解析失败: {e}")
        if attr is None:
            continue
        version = "2.0" if str(rule.get("config_version")) == "2.0" else "1.0"
        entry = {"version": version, "rule": rule, "attr": attr, "v1": None, "error": None}
        try:
            if version == "2.0":
                validate_config_v2(rule, attr)
            else:
                entry["v1"] = build_runtime_config(attr, rule)
        except CallieError as e:
            entry["error"] = str(e)
        cfgs[key] = entry
    return cfgs


# ─────────────────────────────────────────────────────────────────────────────
# 5) 生成 AZ 列（纯函数，完全由 cfg 驱动）
# ─────────────────────────────────────────────────────────────────────────────

def _kw_match(keywords, text_low):
    """方案1：关键词组合匹配。keywords 全部按词边界出现在 text_low 中即命中。"""
    for kw in keywords:
        if not re.search(r'(?:\b|_)' + re.escape(kw.lower()) + r'(?:\b|_)', text_low):
            return False
    return True


def _normalize_style(style_name, raw, cfg):
    """样式值归一：先精确（方案3）再关键词组合（方案1）。返回 (标准值, 是否命中)。"""
    if not raw:
        return raw, True
    if raw in cfg.get("style_alias_exact", {}):
        return cfg["style_alias_exact"][raw], True
    low = raw.lower()
    for r in cfg.get("style_alias_rules", []):
        if _kw_match(r.get("keywords", []), low):
            return r["standard"], True
    return raw, False


def _default_output_order(roles, style_name, char_name, hair_style_field,
                          beard_style_field, attr_groups, routed_style_keys,
                          routed_beard_keys):
    """未显式配置 output_order 时，按【属性表组顺序】给出确定性固定输出顺序。"""
    order = [style_name, "__fixed_prefix__", char_name]
    for role in roles:
        if hair_style_field:
            order.append(f"{role}'s {hair_style_field}")
        if beard_style_field:
            order.append(f"{role}'s {beard_style_field}")
    skip = {style_name, char_name} | set(routed_style_keys) | set(routed_beard_keys)
    for g in attr_groups:
        if g["name"] in skip:
            continue
        order.append(g["name"])
    return order


def generate_callie(j_text, cfg, collect_unmatched=False):
    """
    输入：J 列文本 + 运行时配置（见 build_runtime_config）
    输出：AZ 列多行文本（不含末尾换行）。

    关键特性（用户硬性要求）：
      * 输出顺序由 cfg["output_order"] 决定，与 J 列输入顺序无关（难点1）。
      * 样式值经归一化（精确 + 关键词组合），与输入写法无关（难点2）。
      * 未知定制项Key / 未知角色 / 未知发色 -> 抛 CallieError。

    collect_unmatched=True 时返回 (文本, [未命中归一化的字段列表])，供 UI 自学习。
    """
    data = parse_j(j_text)
    if not data:
        return "" if not collect_unmatched else ("", [])

    # 1) 过滤键（大小写不敏感）
    filtered = OrderedDict()
    for k, v in data.items():
        if k.lower() in cfg["filter_keys"]:
            continue
        filtered[k] = v

    # 2) 角色
    char_name = cfg["character_field"]
    character = filtered.get(char_name)
    if character not in cfg["roles"]:
        raise CallieError(f"未知角色值: {character!r}")
    role = character

    # 3) 未知 Key 校验（路由样式/胡子键已放行）
    for k in filtered:
        if k in (char_name, cfg["style_field"]):
            continue
        if k not in cfg["valid_keys"]:
            raise CallieError(f"未知定制项Key: {k!r}")

    # 4) 先把所有字段都解析并路由好，存进 out[字段名] = [行...]
    out = OrderedDict()
    style_name = cfg["style_field"]
    raw_style = filtered.get(style_name, "")
    norm_style, _matched = _normalize_style(style_name, raw_style, cfg)
    out[style_name] = [f"{cfg['style_gid']}|{style_name}:|{norm_style}"]

    out[char_name] = [f"{cfg['char_gid']}|{char_name}:|{character}"]

    hair_style_field = cfg["hair_style_field"]
    beard_style_field = cfg["beard_style_field"]
    if hair_style_field:
        key = f"{role}'s {hair_style_field}"
        if key in filtered:
            hc_key = f"{role}'s {cfg.get('_color_field_hair')}"
            hc = filtered.get(hc_key, "")
            rt = cfg["routing"].get(role, {})
            if hc in rt:
                out[key] = [f"{rt[hc]}|{key}:|{filtered.get(key, '')}"]
            else:
                nh = cfg["no_hair"].get(role)
                if nh and hc == nh[1]:
                    out[key] = [f"{nh[0]}|{cfg['_no_hair_field']}:|{hc}"]
                else:
                    raise CallieError(f"未知发色值: {hc!r}，无法匹配发型ID")
    if beard_style_field:
        key = f"{role}'s {beard_style_field}"
        if key in filtered:
            bc_key = f"{role}'s {cfg.get('_color_field_beard')}"
            bc = filtered.get(bc_key, "")
            rt = cfg["beard_routing"].get(role, {})
            if bc in rt:
                out[key] = [f"{rt[bc]}|{key}:|{filtered.get(key, '')}"]
            else:
                nb = cfg["no_beard"].get(role)
                if nb and bc == nb[1]:
                    out[key] = [f"{nb[0]}|{cfg['_no_beard_field']}:|{bc}"]
                else:
                    raise CallieError(f"未知胡子颜色值: {bc!r}，无法匹配胡子ID")

    # 普通 / 共享字段
    for k, v in filtered.items():
        if k in (char_name, style_name):
            continue
        if hair_style_field and k == f"{role}'s {hair_style_field}":
            continue
        if beard_style_field and k == f"{role}'s {beard_style_field}":
            continue
        if k not in cfg["groups"]:
            continue
        if not v:
            continue
        out[k] = [f"{cfg['groups'][k]}|{k}:|{v}"]

    # 5) 按 output_order 固定顺序输出（与输入顺序无关）
    order = cfg.get("output_order") or []
    lines = []
    emitted = set()
    for token in order:
        if token == "__fixed_prefix__":
            lines.extend(cfg["fixed_prefix"])
            emitted.add(token)
            continue
        if token == style_name:
            if style_name in out:
                lines.extend(out[style_name])
                emitted.add(style_name)
            continue
        if token == char_name:
            if char_name in out:
                lines.extend(out[char_name])
                emitted.add(char_name)
            continue
        if "{role}" in token:
            rk = token.replace("{role}", role)
            if rk in out:
                lines.extend(out[rk])
                emitted.add(rk)
            continue
        if token in out:
            lines.extend(out[token])
            emitted.add(token)
    # 兜底：未在 output_order 中出现的字段，按属性表组顺序补齐
    for gname in cfg["groups"]:
        if gname in out and gname not in emitted:
            lines.extend(out[gname])
            emitted.add(gname)
    for k in out:
        if k not in emitted:
            lines.extend(out[k])
            emitted.add(k)

    result = "\n".join(lines)
    if collect_unmatched:
        unmatched = []
        if raw_style and not _matched:
            unmatched.append({"field": style_name, "raw": raw_style})
        return result, unmatched
    return result


# ─────────────────────────────────────────────────────────────────────────────
# 5b) v2 配置：显式「J键 → AZ」翻译层（rule.json 含 config_version == "2.0"）
#
# 设计（用户决策：C 混合 + 三类来源 + 先J后补 + 零写死）：
#   * 手写 ID，但启动时用属性表校验合法性（防 105 写成 150）。
#   * 三类来源：固定行 / 固定字段补全 / 客户定制项（J 键→AZ 行）。
#   * 固定字段补全：J 列有则用 J 值，没有才用默认值。
#   * 客户定制项支持：值映射 / dynamic 路由 / 字段名模板 / 键归一化。
# ─────────────────────────────────────────────────────────────────────────────

def _fill_template(tmpl, value_by_field):
    """把模板里 [字段名] 替换为对应字段当前值（如 Name([Color]) -> Name(Black)）。"""
    def _repl(m):
        return str(value_by_field.get(m.group(1), ""))
    return re.sub(r"\[([^\]]+)\]", _repl, tmpl)


# 占位符（如 +[170|Color:|[Color]] 中的 [Color]）—— 从 J 列取值；缺失记警告并留空
_PLACEHOLDER_RE = re.compile(r"\[([^\]]+)\]")


def _sub_placeholders(raw, norm, warnings):
    """把 raw 中的 [JKey] 替换为 J 列对应值；缺失则输出空值并记录一条警告。"""
    def _rep(m):
        key = m.group(1).strip()
        val = norm.get(key, "")
        if val == "":
            warnings.append({"field": key, "raw": ""})
        return val
    return _PLACEHOLDER_RE.sub(_rep, raw)


def _when_ok(when, norm):
    """所有条件（AND）都满足才执行。

    when 可为：
      * 单 dict（兼容旧数据 / 单条件）：{"field":..,"value":..}
      * dict 列表（& 连接的多条件）：[{field,value}, ...]  —— 全部命中才 True
    空 / None -> 视为无守卫，直接 True。
    """
    if not when:
        return True
    conds = when if isinstance(when, list) else [when]
    for c in conds:
        if norm.get(c["field"], "") != c["value"]:
            return False
    return True


def _resolve_gid_by_template(attr, name, active_templates):
    """按活跃模板查字段的 gid：性别模板（非 public）优先，public 补充，最后兜底第一个。

    active_templates：集合，如 {"Woman", "public"}。模板字段值来自 J 列（rule.template_field）。
    """
    groups = [g for g in (attr.get("groups", []) or []) if g["name"] == name]
    if not groups:
        return None
    # 性别模板优先
    for g in groups:
        if g.get("tmpl", "") in active_templates and g.get("tmpl", "") != "public":
            return str(g["gid"])
    # 回退 public
    for g in groups:
        if g.get("tmpl", "") == "public":
            return str(g["gid"])
    # 兜底：第一个（模板值在属性表无对应 tmpl 时）
    return str(groups[0]["gid"])


def validate_config_v2(rule, attr):
    """校验 v2 配置里出现的所有 ID 都存在于属性表（C 混合模式的启动期校验）。"""
    valid_ids = set(str(g["gid"]) for g in attr.get("groups", []))
    problems = []

    def _check(idv, where):
        if idv in ("dynamic", "template"):
            return
        if str(idv) not in valid_ids:
            problems.append(f"{where} 的 ID {idv!r} 不存在于属性表")

    for f in rule.get("fixed_fields", []) or []:
        _check(f.get("id"), f"fixed_fields[{f.get('field')}]")
    for m in rule.get("field_mapping", []) or []:
        _check(m.get("id"), f"field_mapping[{m.get('j_key')}]")
    for r in rule.get("conditional_routing", []) or []:
        tgt = r.get("target_field")
        if r.get("fixed_id") is not None:
            _check(r["fixed_id"], f"conditional_routing[{tgt}]/fixed_id")
            continue
        for dep_val, tid in (r.get("rules") or {}).items():
            _check(tid, f"conditional_routing[{tgt}]/{dep_val}")
    if problems:
        raise CallieError("；".join(problems))


def generate_callie_v2(j_text, rule, attr=None, collect_unmatched=False):
    """
    v2 引擎：输入 J 列文本 + v2 rule.json（+ 可选 attr 用于校验），输出 AZ 列文本。

    处理流程（与用户定义的「先J后补」一致）：
      1. 解析 J 为字典；
      2. key_normalize 归一键名（保留全部，含 ignore_keys，供模板/路由引用，不立即删除）；
      3. fixed_fields 补全（J 有则用 J 值，无/空则取默认值，ID 用配置）；
      4. field_mapping 逐字段翻译（值映射 value_maps / dynamic 路由 conditional_routing /
         字段名模板 field_template）；
      5. 按 output_order 组装（fixed_lines 永远最前；ignore_keys 的字段不单独成行）。

    collect_unmatched=True 时返回 (文本, 未匹配字段列表)。
    """
    if attr is not None:
        validate_config_v2(rule, attr)

    raw = parse_j(j_text)
    if not raw:
        return "" if not collect_unmatched else ("", [])

    # 键归一化：J 键 -> 规范字段名（供 field_mapping / 路由 / 模板引用）
    kn = rule.get("key_normalize", {}) or {}
    norm = OrderedDict()
    for k, v in raw.items():
        norm[kn.get(k, k)] = v

    # 模板选择：读 J 的模板字段值 → 活跃模板集（值 + public）。
    # 仅当声明了 @template 且属性表可用时生效；字段 gid 查询按此模板集过滤。
    active_templates = None
    template_field = rule.get("template_field")
    if attr is not None and template_field:
        tv = norm.get(template_field, "").strip()
        if tv:
            active_templates = {tv, "public"}

    # 忽略字段：不输出自身行，但值保留在 norm 供模板 / 路由引用。
    # 按 key_normalize 对齐，确保「原生 J 键 → 规范名」与 field_mapping 比较一致。
    ignore = {kn.get(k, k) for k in (rule.get("ignore_keys", []) or [])}
    value_maps = rule.get("value_maps", {}) or {}
    routing_by_target = {}
    for r in (rule.get("conditional_routing", []) or []):
        routing_by_target.setdefault(r["target_field"], []).append(r)
    unmatched = []

    # 属性表 name->gid 映射（用于「隐式字段映射」：D 未显式列出的 J 字段自动查表补 ID）
    attr_by_name = {}
    if attr is not None:
        for g in attr.get("groups", []):
            attr_by_name.setdefault(g["name"], str(g["gid"]))

    # conditional_routing 的目标字段必须产出 AZ 行（dynamic 路由），即使 D 未显式列出 [字段]
    field_mapping_all = list(rule.get("field_mapping", []) or [])
    _seen_az = {m["az_field"] for m in field_mapping_all}
    _fixed_field_names = {f["field"] for f in (rule.get("fixed_fields", []) or [])}
    for r in (rule.get("conditional_routing", []) or []):
        tf = r["target_field"]
        if tf in _fixed_field_names:
            continue  # 同时是固定字段的路由目标：由下方 fixed_fields 循环处理条件覆盖
        if tf not in _seen_az:
            field_mapping_all.append({"j_key": tf, "id": "dynamic", "az_field": tf})
            _seen_az.add(tf)

    # 供条件守卫 / 路由依赖 使用的 norm（含固定字段默认值），使 +[Age:Young] 的默认也能参与条件判断
    norm_cond = OrderedDict(norm)
    for f in (rule.get("fixed_fields", []) or []):
        if norm_cond.get(f["field"], "") == "":
            norm_cond[f["field"]] = f.get("default", "")

    # 固定字段补全：J 有则用 J 值，无/空则取默认值
    fixed_out = OrderedDict()
    for f in rule.get("fixed_fields", []) or []:
        field = f["field"]
        value = norm.get(field, "")
        if not value:
            value = f.get("default", "")
        # 条件覆盖：该固定字段同时是 conditional_routing 目标时，
        # 若 when 守卫全部满足且依赖值命中，则改用路由出的 ID（否则回退基础 ID）。
        r_list = routing_by_target.get(field)
        if r_list:
            chosen = None
            for r in r_list:
                if r.get("fixed_id") is not None:
                    if _when_ok(r.get("when"), norm_cond):
                        chosen = r
                        break
                elif _when_ok(r.get("when"), norm_cond) and \
                        norm_cond.get(r["dependency_field"], "") in (r.get("rules") or {}):
                    chosen = r
                    break
            if chosen is not None:
                if chosen.get("fixed_id") is not None:
                    fixed_out[field] = (str(chosen["fixed_id"]), value)
                    continue
                tid = (chosen["rules"]).get(norm_cond.get(chosen["dependency_field"], ""))
                if tid is not None:
                    fixed_out[field] = (str(tid), value)
                    continue
        fid = f["id"]
        if fid == "template" and active_templates and attr is not None:
            fid = _resolve_gid_by_template(attr, field, active_templates) or fid
        fixed_out[field] = (str(fid), value)

    # 客户定制项逐字段翻译（含 augmented 的 dynamic 路由目标字段）
    map_out = OrderedDict()
    for m in field_mapping_all:
        jk = m["j_key"]
        az = m["az_field"]
        if jk in ignore:
            continue  # 忽略字段：不输出自身行（值已保留在 norm 供模板 / 路由引用）
        rv = norm.get(jk, "")
        if not rv:
            continue  # 无值则不输出该行（固定字段由 fixed_fields 处理默认值）
        if az in value_maps and rv in value_maps[az]:
            rv = value_maps[az][rv]
        mid = m["id"]
        if mid == "template" and active_templates and attr is not None:
            mid = _resolve_gid_by_template(attr, az, active_templates) or mid
        if mid == "dynamic":
            r_list = routing_by_target.get(az)
            if not r_list:
                raise CallieError(f"字段 {az!r} 标记为 dynamic，但无对应 conditional_routing 配置")
            # 守卫筛选：先匹配 when 守卫（& 连接的多条件需全部命中）；有守卫但无一命中 → 跳过该行
            guarded = [r for r in r_list if r.get("when")]
            if guarded:
                chosen = None
                for r in guarded:
                    if _when_ok(r.get("when"), norm_cond):
                        chosen = r
                        break
                if chosen is None:
                    continue  # 条件均不匹配 → 跳过该行
            else:
                chosen = r_list[0]
            r = chosen
            if r.get("fixed_id") is not None:
                # 固定ID模式：ID 由 when 守卫直接给出，值取 J 的 target_field 内容（rv 已是）
                mid = str(r["fixed_id"])
            else:
                dep = norm_cond.get(r["dependency_field"], "")
                if not dep:
                    continue  # 无依赖值，无法路由，跳过该行
                tid = (r.get("rules") or {}).get(dep)
                if tid is None:
                    raise CallieError(
                        f"{r['dependency_field']}={dep!r} 无法匹配 {az} 的 ID（请检查 conditional_routing 规则）")
                mid = str(tid)
        label = az
        tmpl = m.get("field_template")
        if tmpl:
            label = _fill_template(tmpl, norm)
        map_out[az] = f"{mid}|{label}:|{rv}"

    # 组装：固定行（字面 / 含占位符）最前；占位符从 J 取值，缺失记警告并输出空值
    warnings = []
    lines = [_sub_placeholders(l, norm, warnings)
             for l in (rule.get("fixed_lines", []) or [])]
    # 条件固定行：when 守卫命中才输出（同样支持占位符）
    for cfl in (rule.get("conditional_fixed_lines", []) or []):
        when = cfl.get("when")
        if when and not _when_ok(when, norm_cond):
            continue
        lines.append(_sub_placeholders(cfl["raw"], norm, warnings))

    order = rule.get("output_order", []) or []
    if not order:
        order = [f["field"] for f in (rule.get("fixed_fields", []) or [])] + \
                [m["az_field"] for m in (rule.get("field_mapping", []) or [])]
    emitted = set()
    for name in order:
        if name in fixed_out and name not in emitted:
            fid, fval = fixed_out[name]
            lines.append(f"{fid}|{name}:|{fval}")
            emitted.add(name)
        elif name in map_out and name not in emitted:
            lines.append(map_out[name])
            emitted.add(name)
    # 兜底：未在 output_order 中出现的字段按声明顺序补在末尾（防丢数据）
    for name, (fid, fval) in fixed_out.items():
        if name not in emitted:
            lines.append(f"{fid}|{name}:|{fval}")
            emitted.add(name)
    for name, line in map_out.items():
        if name not in emitted:
            lines.append(line)
            emitted.add(name)

    # 隐式字段映射：J 列中出现、但未在 D 显式列出的字段，自动查属性表补 ID 输出 AZ 行。
    # 已在 field_mapping / fixed_fields / ignore 处理的字段跳过；
    # 若属性表中找不到该字段 -> 不输出，留待下方 unmatched 收集器报「部分未匹配」错误（防静默失败）。
    if attr_by_name:
        handled_fields = {m["j_key"] for m in field_mapping_all} | \
                         {f["field"] for f in (rule.get("fixed_fields", []) or [])}
        # 固定行（fixed_lines + conditional_fixed_lines 字面行）已占用的 gid 集合。
        # 若某字段的 gid 已被固定行占用（如 Style gid=1 被 +[1|:10001|] 占用），
        # 则隐式映射不再输出该字段（否则会冗余输出 1|Style:|Vertical）。
        fixed_line_gids = set()
        for l in (rule.get("fixed_lines", []) or []):
            gid = str(l).split("|", 1)[0].strip()
            if gid:
                fixed_line_gids.add(gid)
        for cfl in (rule.get("conditional_fixed_lines", []) or []):
            raw = str(cfl.get("raw", "") or "")
            gid = raw.split("|", 1)[0].strip()
            if gid:
                fixed_line_gids.add(gid)
        for key in norm:
            if key in handled_fields or key in ignore:
                continue
            if active_templates and attr is not None:
                gid = _resolve_gid_by_template(attr, key, active_templates)
            else:
                gid = attr_by_name.get(key)
            if gid is None:
                continue
            if str(gid) in fixed_line_gids:
                continue  # gid 已被固定行占用，不隐式输出（避免冗余/重复）
            val = norm.get(key, "")
            lines.append(f"{gid}|{key}:|{val}")
            emitted.add(key)

    # 已写入 lines 的固定行 / 条件固定行 / 隐式映射行：若其 label 对应 J 中的某字段，
    # 视为「该字段已输出」，避免条件固定行（如 ? [Background Style:Crayon],+[170|Color:|[Color]]）
    # 实际已输出却仍被下方「未匹配」收集器误报为部分未匹配。
    # 例：PL251368 中 Background Style=Crayon 时 Color 由条件固定行输出到 AZ，
    # 但隐式映射因 gid=170 已被固定行占用而跳过，导致 Color 被误报（2026-08-21 踩坑）。
    for ln in lines:
        _parts = str(ln).split("|", 2)
        if len(_parts) >= 2:
            _label = _parts[1].strip()
            if _label.endswith(":"):
                _label = _label[:-1].strip()
            if _label and _label in norm:
                emitted.add(_label)

    # 未匹配字段收集（J 中出现但不属于任何 mapping/ignore/fixed/路由依赖的键）
    mapped_keys = set()
    for m in (rule.get("field_mapping", []) or []):
        mapped_keys.add(m["j_key"])
    for f in (rule.get("fixed_fields", []) or []):
        mapped_keys.add(f["field"])
    mapped_keys |= ignore
    mapped_keys |= set((kn or {}).keys())
    for cfl in (rule.get("conditional_fixed_lines", []) or []):
        w = cfl.get("when")
        if w:
            for c in (w if isinstance(w, list) else [w]):
                mapped_keys.add(c["field"])
    for r in (rule.get("conditional_routing", []) or []):
        mapped_keys.add(r["dependency_field"])
        w = r.get("when")
        if w:
            for c in (w if isinstance(w, list) else [w]):
                mapped_keys.add(c["field"])
    for k in norm:
        if k not in mapped_keys and k not in emitted:
            unmatched.append({"field": k, "raw": norm[k]})

    result = "\n".join(lines)
    if collect_unmatched:
        return result, unmatched, warnings
    return result


def generate_callie_dispatch(j_text, sku_cfg, collect_unmatched=False):
    """按 sku_cfg（load_sku_configs 的 value）的版本分发到 v1 / v2 引擎。
    sku_cfg = {"version","rule","attr","v1","error"}。"""
    if sku_cfg.get("error"):
        raise CallieError(f"配置校验失败：{sku_cfg['error']}")
    if sku_cfg.get("version") == "2.0":
        if collect_unmatched:
            text, unm, warn = generate_callie_v2(
                j_text, sku_cfg["rule"], sku_cfg.get("attr"), collect_unmatched=True)
            return text, unm, warn
        return generate_callie_v2(
            j_text, sku_cfg["rule"], sku_cfg.get("attr"), collect_unmatched=False)
    return generate_callie(j_text, sku_cfg["v1"], collect_unmatched=collect_unmatched)


# ─────────────────────────────────────────────────────────────────────────────
# 便捷：从 attribute.xlsx + rule.json 直接构建单 SKU 配置
# ─────────────────────────────────────────────────────────────────────────────

def build_sku_config(attribute_source, rule):
    """
    attribute_source: attribute.xlsx 路径，或已编译的 attribute_config.json 路径。
    rule: rule.json 路径 或 已加载的 dict。
    返回运行时配置。
    """
    if isinstance(rule, str):
        rule = load_rule_config(rule)
    if attribute_source.endswith(".json"):
        with open(attribute_source, encoding="utf-8") as f:
            attr = json.load(f)
    else:
        attr = compile_attribute(attribute_source)
    return build_runtime_config(attr, rule)


DEFAULT_CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "callie_config.json")


def generate_callie_default(j_text, json_path=DEFAULT_CONFIG_PATH):
    """便捷封装：直接吃一份已合并好的配置 JSON（如有需要）。"""
    with open(json_path, encoding="utf-8") as f:
        cfg = json.load(f)
    return generate_callie(j_text, cfg)


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        out = save_attribute_config(sys.argv[1], DEFAULT_CONFIG_PATH.replace(".json", "_attr.json"))
        print("属性配置已写入:", DEFAULT_CONFIG_PATH.replace(".json", "_attr.json"))
        print("属性组数量:", len(out["groups"]))
    else:
        print("用法: python callie_generator.py <attribute.xlsx>")


# ─────────────────────────────────────────────────────────────────────────────
# J 列标准化前置系统（B 系统「阶段 0」）
# ─────────────────────────────────────────────────────────────────────────────
#
# 动机：B 系统（callie 定制项生成器）在实际使用中，J 列（临时列）数据格式
# 存在大量不一致（组合字段、字段名多余/错误、重复字段、大小写/空格/描述杂质），
# 直接拖慢并干扰规则匹配与属性表查询。标准化前置在「规则执行之前」先把 J 列
# 洗成「字段名规范、无组合字段、无杂质」的标准格式，再交给规则引擎。
#
# 规则语法（写在翻译模板「标准化规则」列，单单元格内换行写多条）：
#     rule|原始行|目标行1|目标行2|...;
#   - `rule` 前缀（大小写不敏感）标记这是一条标准化规则。
#   - 每条规则行尾可带 `;`（可选结束符，解析时自动忽略，与 B 系统映射规则习惯一致）。
#   - 原始行：J 列中某一行的【整行精确匹配】内容（匹配前对 J 每行做 strip）。
#   - 目标行 1..N：标准化后展开成的一到多行；目标行留空（rule|原始行|）→ 删除该行。
#   - 原始行/目标行不得含 `|`（用作字段分隔符）。
#
# 执行语义：单轮、非链式、按规则书写顺序；一条原始行命中第一条能匹配的规则后即替换，
# 不再参与后续规则匹配（避免循环替换）；未匹配行原样保留。

def parse_std_rules(std_text):
    """把「标准化规则」列文本解析成 [(raw, [targets]), ...]。

    raw  = 原始行（已 strip，用于精确匹配）
    targets = 目标行列表（可能为空 = 删除该行）

    返回空列表表示无规则（normalize_j 会原样返回）。
    """
    rules = []
    if not std_text:
        return rules
    for line in str(std_text).splitlines():
        s = line.strip()
        if not s:
            continue
        if not s.lower().startswith("rule|"):  # 仅接受 `rule|` 前缀；其余（空/注释/映射语法）忽略
            continue
        body = s[len("rule|"):]
        if body.endswith(";"):  # 兼容行尾可选结束符
            body = body[:-1]
        parts = body.split("|")
        if len(parts) < 2:
            continue  # 格式残缺，跳过
        raw = parts[0].strip()
        if not raw:
            continue
        # 过滤空目标：`rule|原始行|` 末尾的 `|` 会 split 出空串，须剔除，
        # 使「目标行留空 = 删除该行」语义正确（否则会被当成展开成空行）。
        targets = [t.strip() for t in parts[1:] if t.strip()]
        rules.append((raw, targets))
    return rules


def normalize_j(j_text, std_rules):
    """对 J 列文本执行阶段 0 标准化，返回标准化后的 J 文本。

    std_rules: parse_std_rules 返回的列表。为空则原样返回。
    """
    if not std_rules:
        return j_text
    # 同原始行取第一条规则（书写顺序优先）
    rule_map = {}
    for raw, targets in std_rules:
        if raw not in rule_map:
            rule_map[raw] = targets

    out_lines = []
    for line in str(j_text).splitlines():
        stripped = line.strip()
        if stripped in rule_map:
            targets = rule_map[stripped]
            if targets:  # 一对多展开
                out_lines.extend(targets)
            # 目标为空 = 删除该行（不追加）
        else:
            out_lines.append(line)  # 未匹配：原样保留
    return "\n".join(out_lines)
